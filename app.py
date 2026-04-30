import os
import re
import io
import html
import uuid
import json
import base64
import urllib.parse
import urllib.request
import urllib.error
from datetime import datetime, date, timedelta
from pathlib import Path

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openai import OpenAI
from supabase_config import supabase

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except Exception:
    BARCODE_AVAILABLE = False

# ============================================================
# BRADAFIN MVP V1
# Control financiero inteligente para microempresas
# Streamlit + Supabase + Gemini/OpenAI compatible
# ============================================================

BRADAFIN_LOGO_FULL_PATHS = [
    Path("bradafin_logo_horizontal.png"),
    Path("bradafin_logo_completo.png"),
    Path("bradafin_logo.png"),
    Path("logo_bradafin.png"),
]
BRADAFIN_ICON_PATHS = [
    Path("bradafin_icono.png"),
    Path("icono_bradafin.png"),
    Path("bradafin_isotipo.png"),
]
BRADAFIN_LOGO_FULL_PATH = next((p for p in BRADAFIN_LOGO_FULL_PATHS if p.exists()), None)
BRADAFIN_ICON_PATH = next((p for p in BRADAFIN_ICON_PATHS if p.exists()), None)
APP_PAGE_ICON = str(BRADAFIN_ICON_PATH) if BRADAFIN_ICON_PATH else "💼"

st.set_page_config(
    page_title="BradaFin",
    page_icon=APP_PAGE_ICON,
    layout="wide",
    initial_sidebar_state="collapsed",
)

APP_NAME = "BradaFin"
APP_TAGLINE = "Sepa cuánto vende, cuánto gasta y cuánto gana su negocio."
APP_PROMISE = "Control financiero inteligente para microempresas."
CURRENCY = "COP"

DEFAULT_CATEGORIAS = {
    "Venta": ["Ventas generales", "Mostrador", "Domicilio", "WhatsApp", "Servicios"],
    "Gasto operativo": ["Arriendo", "Servicios", "Transporte", "Nómina", "Mercancía", "Mantenimiento", "Marketing", "Otros"],
    "Entrada de caja": ["Capital", "Préstamo", "Ajuste positivo", "Otros"],
    "Salida de caja": ["Retiro dueño", "Ajuste negativo", "Compra menor", "Otros"],
    "Compra inventario": ["Compra mercancía", "Reposición", "Proveedor", "Otros"],
}

TIPOS_MOVIMIENTO = ["Venta", "Gasto operativo", "Entrada de caja", "Salida de caja", "Compra inventario"]
TIPOS_CUENTA = ["Por cobrar", "Por pagar"]
ESTADOS_CUENTA = ["pendiente", "abonada", "vencida", "pagada"]
METODOS_PAGO = ["Efectivo", "Transferencia", "Nequi", "Daviplata", "Tarjeta", "Crédito", "Otro"]

def render_logo_header(width_full=300, width_icon=82, show_name_fallback=True):
    if BRADAFIN_LOGO_FULL_PATH:
        st.image(str(BRADAFIN_LOGO_FULL_PATH), width=width_full)
    elif BRADAFIN_ICON_PATH:
        st.image(str(BRADAFIN_ICON_PATH), width=width_icon)
        if show_name_fallback:
            st.markdown("### BradaFin")
    elif show_name_fallback:
        st.markdown("### BradaFin")


def render_logo_sidebar(width_full=190, width_icon=70, show_name_fallback=True):
    if BRADAFIN_LOGO_FULL_PATH:
        st.image(str(BRADAFIN_LOGO_FULL_PATH), width=width_full)
    elif BRADAFIN_ICON_PATH:
        st.image(str(BRADAFIN_ICON_PATH), width=width_icon)
        if show_name_fallback:
            st.markdown("### BradaFin")
    elif show_name_fallback:
        st.markdown("### BradaFin")

# ============================================================
# CONFIG
# ============================================================

def leer_config(clave, default=""):
    try:
        valor = st.secrets.get(clave)
    except Exception:
        valor = None
    if valor in (None, ""):
        valor = os.getenv(clave, default)
    return valor if valor is not None else default


def leer_float(clave, default):
    try:
        return float(str(leer_config(clave, default)).strip())
    except Exception:
        return float(default)

GEMINI_API_KEY = leer_config("GEMINI_API_KEY", "")
BRADAFIN_IA_TIMEOUT = leer_float("BRADAFIN_IA_TIMEOUT", 22)
GEMINI_MODEL_CANDIDATES = [
    str(leer_config("GEMINI_MODEL_PRIMARY", "gemini-2.5-flash")).strip(),
    str(leer_config("GEMINI_MODEL_FALLBACK_1", "gemini-2.0-flash")).strip(),
    str(leer_config("GEMINI_MODEL_FALLBACK_2", "gemini-1.5-flash")).strip(),
]
GEMINI_MODEL_CANDIDATES = [m for m in GEMINI_MODEL_CANDIDATES if m]

openai_client = None
if GEMINI_API_KEY:
    try:
        openai_client = OpenAI(
            api_key=GEMINI_API_KEY,
            base_url="https://generativelanguage.googleapis.com/v1beta/openai/",
            timeout=BRADAFIN_IA_TIMEOUT,
            max_retries=0,
        )
    except Exception:
        openai_client = None

# ============================================================
# WHATSAPP AUTOMÁTICO
# ============================================================
# BradaFin puede enviar WhatsApp automático si configuras Twilio o Meta Cloud API.
# Si no hay credenciales, la app muestra el mensaje premium listo para abrir en WhatsApp.
BRADAFIN_AUTO_WHATSAPP = str(leer_config("BRADAFIN_AUTO_WHATSAPP", "false")).strip().lower() in ["1", "true", "yes", "si", "sí", "on"]
WHATSAPP_PROVIDER = str(leer_config("WHATSAPP_PROVIDER", "manual")).strip().lower()
WHATSAPP_COUNTRY_CODE = str(leer_config("WHATSAPP_COUNTRY_CODE", "57")).strip()

# Twilio WhatsApp
TWILIO_ACCOUNT_SID = leer_config("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = leer_config("TWILIO_AUTH_TOKEN", "")
TWILIO_WHATSAPP_FROM = leer_config("TWILIO_WHATSAPP_FROM", "")

# Meta WhatsApp Cloud API
META_WHATSAPP_TOKEN = leer_config("META_WHATSAPP_TOKEN", "")
META_WHATSAPP_PHONE_NUMBER_ID = leer_config("META_WHATSAPP_PHONE_NUMBER_ID", "")

# ============================================================
# ESTILO
# ============================================================

def aplicar_estilo_bradafin():
    st.markdown(
        """
        <style>
        :root {
            --bg: #FBF8EF;
            --bg2: #F4F8F2;
            --surface: #FFFFFF;
            --text: #102019;
            --muted: #4F6258;
            --green: #1F6B4F;
            --green2: #14513D;
            --green3: #2F8F6B;
            --gold: #D4A017;
            --gold2: #F2D16B;
            --gold3: #FFF3C4;
            --red: #C2410C;
            --red2: #EF4444;
            --line: rgba(16,32,25,.12);
            --shadow: 0 18px 40px rgba(20,81,61,.10);
            --shadow2: 0 10px 24px rgba(16,32,25,.07);
        }

        html, body, [class*="css"] {
            color: var(--text) !important;
        }

        .stApp {
            background:
                radial-gradient(circle at top left, rgba(212,160,23,.13), transparent 26%),
                radial-gradient(circle at top right, rgba(31,107,79,.10), transparent 24%),
                linear-gradient(180deg, #FFFCF3 0%, #F6FAF4 100%);
        }

        header[data-testid="stHeader"] {
            background: rgba(255,255,255,.86);
            backdrop-filter: blur(14px);
            border-bottom: 1px solid rgba(16,32,25,.05);
        }

        .block-container {
            max-width: 1360px;
            padding-top: 2.35rem;
            padding-bottom: 4rem;
        }

        .login-brand-spacer {
            height: 1.95rem;
        }

        .login-logo-after {
            height: 1.30rem;
        }

        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #F9F4E6 100%);
            border-right: 1px solid rgba(16,32,25,.08);
        }

        [data-testid="collapsedControl"] {
            position: fixed !important;
            top: .85rem !important;
            left: .85rem !important;
            z-index: 999999 !important;
            background: linear-gradient(135deg, #14513D 0%, #1F6B4F 55%, #D4A017 100%) !important;
            border-radius: 16px !important;
            box-shadow: 0 14px 28px rgba(31,107,79,.20) !important;
        }
        [data-testid="collapsedControl"] button,
        [data-testid="collapsedControl"] svg {
            color: white !important;
            fill: white !important;
        }

        h1, h2, h3, h4, h5, h6,
        p, span, div, label, small,
        .stMarkdown, .stMarkdown p, .stMarkdown span, .stMarkdown div {
            color: var(--text) !important;
        }

        a {
            color: var(--green) !important;
            font-weight: 800 !important;
        }

        /* =================================================== */
        /* FORMULARIOS Y WIDGETS: TEXTO OSCURO EN FONDOS CLAROS */
        /* =================================================== */
        label,
        .stSelectbox label,
        .stTextInput label,
        .stNumberInput label,
        .stDateInput label,
        .stTextArea label,
        .stRadio label,
        .stCheckbox label,
        .stMultiSelect label,
        [data-testid="stWidgetLabel"],
        [data-testid="stWidgetLabel"] p,
        [data-testid="stWidgetLabel"] span {
            color: var(--text) !important;
            font-weight: 850 !important;
        }

        div[data-testid="stRadio"] label,
        div[data-testid="stRadio"] label p,
        div[data-testid="stRadio"] label span,
        div[data-testid="stRadio"] div[role="radiogroup"] label,
        div[data-testid="stRadio"] div[role="radiogroup"] span,
        div[data-testid="stRadio"] div[role="radiogroup"] p {
            color: var(--text) !important;
            font-weight: 900 !important;
            opacity: 1 !important;
        }

        div[data-testid="stRadio"] label:hover p,
        div[data-testid="stRadio"] label:hover span {
            color: var(--green2) !important;
        }

        div[data-testid="stCheckbox"] label,
        div[data-testid="stCheckbox"] label p,
        div[data-testid="stCheckbox"] label span,
        div[data-testid="stCheckbox"] span {
            color: var(--text) !important;
            font-weight: 850 !important;
            opacity: 1 !important;
        }

        .stCaption,
        .stCaption p,
        .stCaption span,
        .caption,
        [data-testid="stCaptionContainer"],
        [data-testid="stCaptionContainer"] p,
        [data-testid="stCaptionContainer"] span {
            color: var(--muted) !important;
            opacity: 1 !important;
        }

        .stTextInput input,
        .stNumberInput input,
        .stDateInput input,
        textarea,
        .stTextArea textarea,
        .stSelectbox div[data-baseweb="select"] > div,
        .stMultiSelect div[data-baseweb="select"] > div {
            border-radius: 16px !important;
            min-height: 48px !important;
            background: #FFFFFF !important;
            color: var(--text) !important;
            border: 1px solid rgba(96,114,102,.28) !important;
            box-shadow: 0 6px 12px rgba(16,32,25,.035) !important;
        }

        input::placeholder,
        textarea::placeholder {
            color: #718076 !important;
            opacity: 1 !important;
        }

        .stSelectbox div[data-baseweb="select"] span,
        .stSelectbox div[data-baseweb="select"] div,
        .stMultiSelect div[data-baseweb="select"] span,
        .stMultiSelect div[data-baseweb="select"] div {
            color: var(--text) !important;
            opacity: 1 !important;
        }

        /* Menús desplegables: si el fondo es oscuro, el texto queda claro y legible */
        div[data-baseweb="popover"] div[role="listbox"],
        div[data-baseweb="popover"] ul {
            background: #102019 !important;
            border: 1px solid rgba(242,209,107,.26) !important;
            border-radius: 18px !important;
            box-shadow: 0 20px 36px rgba(16,32,25,.28) !important;
            padding: .35rem !important;
        }

        div[data-baseweb="popover"] div[role="option"],
        div[data-baseweb="popover"] div[role="option"] span,
        div[data-baseweb="popover"] li,
        div[data-baseweb="popover"] li span {
            color: #FFFFFF !important;
            font-weight: 850 !important;
            opacity: 1 !important;
        }

        div[data-baseweb="popover"] div[role="option"]:hover,
        div[data-baseweb="popover"] li:hover {
            background: linear-gradient(135deg, rgba(31,107,79,.95), rgba(212,160,23,.85)) !important;
            border-radius: 14px !important;
        }

        div[data-baseweb="popover"] [aria-selected="true"] {
            background: linear-gradient(135deg, #1F6B4F 0%, #D4A017 100%) !important;
            border-radius: 14px !important;
        }

        /* Number input +/- premium */
        [data-testid="stNumberInput"] button {
            background: #102019 !important;
            color: #FFFFFF !important;
            border-color: #102019 !important;
        }
        [data-testid="stNumberInput"] button svg {
            color: #FFFFFF !important;
            fill: #FFFFFF !important;
        }

        /* =================================================== */
        /* BOTONES PREMIUM */
        /* =================================================== */
        .stButton > button,
        .stDownloadButton > button,
        button[data-testid="baseButton-secondary"] {
            border-radius: 18px !important;
            min-height: 50px !important;
            font-weight: 950 !important;
            letter-spacing: -.01em !important;
            border: 1px solid rgba(31,107,79,.18) !important;
            color: var(--green2) !important;
            background:
                linear-gradient(#FFFFFF, #FFFFFF) padding-box,
                linear-gradient(135deg, rgba(31,107,79,.55), rgba(212,160,23,.65)) border-box !important;
            box-shadow: 0 12px 22px rgba(16,32,25,.075) !important;
            transition: all .18s ease !important;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        button[data-testid="baseButton-secondary"]:hover {
            transform: translateY(-1px) !important;
            color: var(--text) !important;
            background:
                linear-gradient(180deg, #FFF9E6, #F3FBF5) padding-box,
                linear-gradient(135deg, #1F6B4F, #D4A017) border-box !important;
            box-shadow: 0 18px 30px rgba(31,107,79,.14) !important;
        }

        .stButton > button[kind="primary"],
        .stDownloadButton > button[kind="primary"],
        button[data-testid="baseButton-primary"] {
            background: linear-gradient(135deg, #102019 0%, #14513D 35%, #1F6B4F 70%, #D4A017 100%) !important;
            color: white !important;
            border: none !important;
            box-shadow: 0 18px 34px rgba(31,107,79,.26) !important;
        }

        .stButton > button[kind="primary"] *,
        .stDownloadButton > button[kind="primary"] *,
        button[data-testid="baseButton-primary"] * {
            color: white !important;
        }

        .stButton > button[kind="primary"]:hover,
        .stDownloadButton > button[kind="primary"]:hover,
        button[data-testid="baseButton-primary"]:hover {
            filter: brightness(1.04) saturate(1.05) !important;
            box-shadow: 0 22px 40px rgba(31,107,79,.30) !important;
        }

        section[data-testid="stSidebar"] .stButton > button {
            background: #FFFFFF !important;
            color: var(--green2) !important;
            border: 1px solid rgba(31,107,79,.18) !important;
        }
        section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
            background: linear-gradient(135deg, #14513D 0%, #1F6B4F 70%, #D4A017 100%) !important;
            color: white !important;
            border: none !important;
        }
        section[data-testid="stSidebar"] .stButton > button[kind="primary"] * {
            color: white !important;
        }

        /* =================================================== */
        /* TARJETAS Y COMPONENTES */
        /* =================================================== */
        .hero-card {
            color: white !important;
            border-radius: 34px;
            padding: 1.4rem 1.55rem;
            margin-bottom: 1rem;
            background:
                radial-gradient(circle at top left, rgba(255,255,255,.22), transparent 30%),
                radial-gradient(circle at bottom right, rgba(242,209,107,.20), transparent 26%),
                linear-gradient(135deg, #102019 0%, #14513D 45%, #1F6B4F 72%, #D4A017 100%);
            box-shadow: 0 30px 60px rgba(31,107,79,.24);
        }
        .hero-card,
        .hero-card * {
            color: white !important;
        }
        .hero-badge {
            display:inline-flex;
            padding:.42rem .8rem;
            border-radius:999px;
            background:rgba(255,255,255,.15);
            border:1px solid rgba(255,255,255,.22);
            font-size:.82rem;
            font-weight:900;
            margin-bottom:.85rem;
        }
        .hero-title {
            color:white !important;
            font-size:2.3rem;
            font-weight:950;
            line-height:1.05;
            margin:0 0 .4rem 0;
        }
        .hero-sub {
            color:rgba(255,255,255,.92) !important;
            line-height:1.6;
            max-width:850px;
        }
        .hero-strip {
            display:grid;
            grid-template-columns: repeat(4,minmax(0,1fr));
            gap:.7rem;
            margin-top:1.1rem;
        }
        .hero-mini {
            padding:.85rem .9rem;
            border-radius:22px;
            background:rgba(255,255,255,.15);
            border:1px solid rgba(255,255,255,.18);
        }
        .hero-mini-label {
            font-size:.72rem;
            font-weight:900;
            opacity:.88;
            margin-bottom:.25rem;
        }
        .hero-mini-value {
            font-weight:950;
            font-size:1.06rem;
            letter-spacing:-.03em;
        }

        .soft-card,
        .kpi-card,
        .alert-card,
        .table-card,
        [data-testid="stForm"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #FBFCF8 100%);
            border: 1px solid rgba(16,32,25,.08);
            border-radius: 26px;
            box-shadow: var(--shadow2);
            padding: 1rem 1.05rem;
            margin-bottom: .9rem;
        }

        .soft-card *,
        .kpi-card *,
        .alert-card *,
        .table-card *,
        [data-testid="stForm"] * {
            color: var(--text) !important;
        }

        .section-title {
            color: var(--text) !important;
            font-size:1.2rem;
            font-weight:950;
            letter-spacing:-.03em;
            margin-bottom:.15rem;
        }
        .section-caption,
        .muted {
            color: var(--muted) !important;
            font-size:.93rem;
            line-height:1.55;
        }

        .kpi-card { min-height:128px; }
        .kpi-label { color:#496052 !important; font-size:.82rem; font-weight:900; margin-bottom:.54rem; }
        .kpi-value { color:var(--text) !important; font-size:1.62rem; font-weight:950; line-height:1.08; letter-spacing:-.04em; }
        .kpi-foot { color:var(--muted) !important; font-size:.82rem; margin-top:.3rem; }
        .kpi-green { background: linear-gradient(180deg, #EAF8F1, #F5FFF8); border-color: rgba(31,107,79,.16); }
        .kpi-gold { background: linear-gradient(180deg, #FFF8DD, #FFFCF2); border-color: rgba(212,160,23,.20); }
        .kpi-red { background: linear-gradient(180deg, #FFF1E8, #FFF8F2); border-color: rgba(194,65,12,.16); }
        .kpi-blue { background: linear-gradient(180deg, #EFF6FF, #F7FBFF); border-color: rgba(37,99,235,.13); }

        .pill {
            display:inline-flex;
            align-items:center;
            gap:.35rem;
            padding:.46rem .78rem;
            border-radius:999px;
            font-size:.83rem;
            font-weight:900;
            border:1px solid rgba(16,32,25,.10);
            background:#FFF;
            color:var(--text) !important;
        }
        .pill-green { background:#EAF8F1; border-color:#BFE5D2; color:#14513D !important; }
        .pill-gold { background:#FFF5CC; border-color:#F2D16B; color:#6B4B00 !important; }
        .pill-red { background:#FFF1E8; border-color:#FDBA74; color:#9A3412 !important; }
        .pill-gray { background:#F8FAFC; border-color:#CBD5E1; color:#334155 !important; }

        .movement-row {
            display:flex;
            justify-content:space-between;
            gap:.85rem;
            align-items:flex-start;
            padding:.8rem 0;
            border-bottom:1px solid rgba(16,32,25,.07);
        }
        .movement-row:last-child { border-bottom:none; }
        .movement-title { font-weight:950; color:var(--text) !important; }
        .movement-sub { color:var(--muted) !important; font-size:.84rem; margin-top:.15rem; }
        .amount-good { color:#1F6B4F !important; font-weight:950; white-space:nowrap; }
        .amount-bad { color:#C2410C !important; font-weight:950; white-space:nowrap; }
        .amount-neutral { color:#6B4B00 !important; font-weight:950; white-space:nowrap; }

        .stTabs [data-baseweb="tab-list"] {
            gap:.45rem;
            background:linear-gradient(180deg,#FFFFFF,#F8FAF3);
            border:1px solid rgba(16,32,25,.08);
            border-radius:22px;
            padding:.42rem;
            margin-bottom:1rem;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius:999px !important;
            color:var(--text) !important;
            font-weight:900 !important;
        }
        .stTabs [data-baseweb="tab"] * {
            color: var(--text) !important;
        }
        .stTabs [aria-selected="true"] {
            background:linear-gradient(135deg,#14513D,#1F6B4F,#D4A017) !important;
            color:white !important;
            box-shadow: 0 10px 18px rgba(31,107,79,.18) !important;
        }
        .stTabs [aria-selected="true"] * { color:white !important; }

        .js-plotly-plot,
        .plotly-graph-div {
            border-radius:22px !important;
            overflow:hidden;
        }

        div[data-testid="stAlert"] * {
            color: var(--text) !important;
        }



        /* =================================================== */
        /* FIX FINAL: HERO CON LETRA CLARA + SIDEBAR HAMBURGUESA */
        /* =================================================== */
        .hero-card,
        .hero-card *,
        .hero-card div,
        .hero-card span,
        .hero-card p,
        .hero-card .hero-title,
        .hero-card .hero-sub,
        .hero-card .hero-badge,
        .hero-card .hero-mini,
        .hero-card .hero-mini-label,
        .hero-card .hero-mini-value {
            color: #FFFFFF !important;
            opacity: 1 !important;
        }

        .hero-card .hero-title {
            color: #FFFFFF !important;
            text-shadow: 0 2px 10px rgba(0,0,0,.18) !important;
        }

        .hero-card .hero-sub {
            color: rgba(255,255,255,.93) !important;
            text-shadow: 0 1px 8px rgba(0,0,0,.14) !important;
        }

        .hero-card .hero-badge {
            color: #FFFFFF !important;
            background: rgba(255,255,255,.18) !important;
            border: 1px solid rgba(255,255,255,.28) !important;
            box-shadow: inset 0 1px 0 rgba(255,255,255,.15) !important;
        }

        .hero-card .hero-mini {
            background: rgba(255,255,255,.16) !important;
            border: 1px solid rgba(255,255,255,.24) !important;
            box-shadow: inset 0 1px 0 rgba(255,255,255,.12) !important;
        }

        /* Fuerza visible el control nativo para abrir/cerrar el sidebar.
           Streamlit puede cambiar el testid segun version, por eso cubrimos varias variantes. */
        [data-testid="collapsedControl"],
        [data-testid="stSidebarCollapsedControl"],
        div[data-testid="collapsedControl"],
        button[aria-label="Open sidebar"],
        button[title="Open sidebar"] {
            display: flex !important;
            visibility: visible !important;
            opacity: 1 !important;
            position: fixed !important;
            top: .85rem !important;
            left: .85rem !important;
            z-index: 999999 !important;
            align-items: center !important;
            justify-content: center !important;
            min-width: 44px !important;
            min-height: 44px !important;
            border-radius: 16px !important;
            background: linear-gradient(135deg, #102019 0%, #14513D 45%, #1F6B4F 75%, #D4A017 100%) !important;
            color: #FFFFFF !important;
            border: 1px solid rgba(255,255,255,.30) !important;
            box-shadow: 0 16px 30px rgba(16,32,25,.28) !important;
        }

        [data-testid="collapsedControl"] button,
        [data-testid="stSidebarCollapsedControl"] button,
        [data-testid="collapsedControl"] svg,
        [data-testid="stSidebarCollapsedControl"] svg,
        button[aria-label="Open sidebar"] svg,
        button[title="Open sidebar"] svg {
            color: #FFFFFF !important;
            fill: #FFFFFF !important;
            stroke: #FFFFFF !important;
            opacity: 1 !important;
        }

        section[data-testid="stSidebar"] button[kind="header"],
        section[data-testid="stSidebar"] button[aria-label="Close sidebar"],
        section[data-testid="stSidebar"] button[title="Close sidebar"] {
            color: #102019 !important;
            opacity: 1 !important;
        }

        section[data-testid="stSidebar"] button[kind="header"] svg,
        section[data-testid="stSidebar"] button[aria-label="Close sidebar"] svg,
        section[data-testid="stSidebar"] button[title="Close sidebar"] svg {
            color: #102019 !important;
            fill: #102019 !important;
            stroke: #102019 !important;
        }



        /* =================================================== */
        /* MEJORA VISUAL FINAL: BOTONES PREMIUM + HAMBURGUESA */
        /* Solo CSS: no cambia funciones, datos, Supabase ni flujo */
        /* =================================================== */

        /* Botones generales mas premium */
        .stButton > button,
        .stDownloadButton > button,
        button[data-testid="baseButton-secondary"] {
            min-height: 52px !important;
            border-radius: 999px !important;
            padding: .7rem 1.05rem !important;
            font-weight: 950 !important;
            letter-spacing: -.015em !important;
            color: #102019 !important;
            border: 1px solid rgba(212,160,23,.42) !important;
            background:
                linear-gradient(180deg, rgba(255,255,255,.98) 0%, rgba(255,250,232,.94) 100%) padding-box,
                linear-gradient(135deg, rgba(31,107,79,.55), rgba(212,160,23,.80)) border-box !important;
            box-shadow:
                0 14px 28px rgba(16,32,25,.08),
                inset 0 1px 0 rgba(255,255,255,.78) !important;
            transition: transform .16s ease, box-shadow .16s ease, filter .16s ease !important;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        button[data-testid="baseButton-secondary"]:hover {
            transform: translateY(-2px) !important;
            color: #102019 !important;
            border-color: rgba(212,160,23,.72) !important;
            background:
                radial-gradient(circle at 18% 10%, rgba(242,209,107,.32), transparent 34%),
                linear-gradient(135deg, #FFFFFF 0%, #FFF6D8 45%, #F2FFF6 100%) !important;
            box-shadow:
                0 22px 40px rgba(31,107,79,.16),
                inset 0 1px 0 rgba(255,255,255,.88) !important;
        }

        .stButton > button:active,
        .stDownloadButton > button:active {
            transform: translateY(0) scale(.99) !important;
        }

        /* Botones principales: verde profundo + dorado */
        .stButton > button[kind="primary"],
        .stDownloadButton > button[kind="primary"],
        button[data-testid="baseButton-primary"] {
            color: #FFFFFF !important;
            border: 1px solid rgba(255,255,255,.16) !important;
            background:
                radial-gradient(circle at 18% 12%, rgba(255,255,255,.22), transparent 30%),
                linear-gradient(135deg, #0B1612 0%, #14513D 38%, #1F6B4F 68%, #D4A017 100%) !important;
            box-shadow:
                0 20px 42px rgba(31,107,79,.26),
                inset 0 1px 0 rgba(255,255,255,.20) !important;
        }

        .stButton > button[kind="primary"] *,
        .stDownloadButton > button[kind="primary"] *,
        button[data-testid="baseButton-primary"] * {
            color: #FFFFFF !important;
        }

        .stButton > button[kind="primary"]:hover,
        .stDownloadButton > button[kind="primary"]:hover,
        button[data-testid="baseButton-primary"]:hover {
            filter: brightness(1.06) saturate(1.08) !important;
            box-shadow:
                0 26px 48px rgba(31,107,79,.31),
                0 0 0 3px rgba(212,160,23,.10),
                inset 0 1px 0 rgba(255,255,255,.25) !important;
        }

        /* Sidebar: botones tipo tarjeta premium */
        section[data-testid="stSidebar"] .stButton > button {
            min-height: 50px !important;
            border-radius: 20px !important;
            justify-content: center !important;
            color: #173B2D !important;
            background:
                linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,252,242,.92)) padding-box,
                linear-gradient(135deg, rgba(31,107,79,.24), rgba(212,160,23,.36)) border-box !important;
            border: 1px solid transparent !important;
            box-shadow:
                0 10px 22px rgba(16,32,25,.055),
                inset 0 1px 0 rgba(255,255,255,.75) !important;
        }

        section[data-testid="stSidebar"] .stButton > button:hover {
            color: #102019 !important;
            background:
                radial-gradient(circle at 15% 10%, rgba(242,209,107,.30), transparent 34%),
                linear-gradient(135deg, #FFFFFF 0%, #FFF8DD 52%, #EEF9F1 100%) !important;
            transform: translateX(2px) !important;
            box-shadow: 0 16px 28px rgba(31,107,79,.12) !important;
        }

        section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
            color: #FFFFFF !important;
            background:
                radial-gradient(circle at 18% 14%, rgba(255,255,255,.22), transparent 28%),
                linear-gradient(135deg, #0B1612 0%, #14513D 42%, #1F6B4F 72%, #D4A017 100%) !important;
            border: 1px solid rgba(255,255,255,.14) !important;
            box-shadow:
                0 18px 34px rgba(31,107,79,.24),
                inset 0 1px 0 rgba(255,255,255,.18) !important;
        }

        section[data-testid="stSidebar"] .stButton > button[kind="primary"] *,
        section[data-testid="stSidebar"] .stButton > button[kind="primary"] p,
        section[data-testid="stSidebar"] .stButton > button[kind="primary"] span {
            color: #FFFFFF !important;
        }

        /* Radio buttons como chips premium, sin cambiar comportamiento */
        div[data-testid="stRadio"] div[role="radiogroup"] {
            gap: .7rem 1rem !important;
            flex-wrap: wrap !important;
            align-items: center !important;
        }

        div[data-testid="stRadio"] label {
            min-height: 42px !important;
            padding: .48rem .82rem !important;
            border-radius: 999px !important;
            background:
                linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,251,238,.92)) padding-box,
                linear-gradient(135deg, rgba(31,107,79,.25), rgba(212,160,23,.38)) border-box !important;
            border: 1px solid transparent !important;
            box-shadow: 0 10px 20px rgba(16,32,25,.055) !important;
            transition: all .16s ease !important;
        }

        div[data-testid="stRadio"] label:hover {
            transform: translateY(-1px) !important;
            box-shadow: 0 16px 26px rgba(31,107,79,.12) !important;
            background:
                radial-gradient(circle at 20% 15%, rgba(242,209,107,.28), transparent 34%),
                linear-gradient(135deg, #FFFFFF 0%, #FFF7D9 50%, #F2FFF6 100%) !important;
        }

        div[data-testid="stRadio"] label:has(input:checked) {
            background:
                radial-gradient(circle at 18% 12%, rgba(255,255,255,.18), transparent 30%),
                linear-gradient(135deg, #102019 0%, #14513D 50%, #D4A017 100%) !important;
            border-color: rgba(255,255,255,.18) !important;
            box-shadow: 0 18px 30px rgba(31,107,79,.22) !important;
        }

        div[data-testid="stRadio"] label:has(input:checked) p,
        div[data-testid="stRadio"] label:has(input:checked) span,
        div[data-testid="stRadio"] label:has(input:checked) div {
            color: #FFFFFF !important;
        }

        /* Checkbox visual premium */
        div[data-testid="stCheckbox"] label {
            width: fit-content !important;
            padding: .55rem .85rem !important;
            border-radius: 18px !important;
            background:
                linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,251,238,.92)) padding-box,
                linear-gradient(135deg, rgba(31,107,79,.25), rgba(212,160,23,.38)) border-box !important;
            border: 1px solid transparent !important;
            box-shadow: 0 10px 20px rgba(16,32,25,.055) !important;
        }

        div[data-testid="stCheckbox"] label:has(input:checked) {
            background:
                linear-gradient(135deg, rgba(31,107,79,.10), rgba(212,160,23,.18)) padding-box,
                linear-gradient(135deg, #1F6B4F, #D4A017) border-box !important;
            box-shadow: 0 14px 24px rgba(31,107,79,.12) !important;
        }

        /* Boton hamburguesa tipo medallon premium a la izquierda */
        [data-testid="collapsedControl"],
        [data-testid="stSidebarCollapsedControl"],
        div[data-testid="collapsedControl"],
        button[aria-label="Open sidebar"],
        button[title="Open sidebar"] {
            width: 56px !important;
            height: 56px !important;
            min-width: 56px !important;
            min-height: 56px !important;
            border-radius: 999px !important;
            position: fixed !important;
            top: 1rem !important;
            left: 1rem !important;
            z-index: 999999 !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            background:
                radial-gradient(circle at 30% 24%, #F4E2C7 0%, #C6A27C 42%, #9B7554 72%, #6F513D 100%) !important;
            border: 1px solid rgba(255,255,255,.38) !important;
            box-shadow:
                0 18px 36px rgba(111,81,61,.33),
                0 0 0 7px rgba(212,160,23,.08),
                inset 0 1px 0 rgba(255,255,255,.35) !important;
            overflow: hidden !important;
        }

        [data-testid="collapsedControl"]::before,
        [data-testid="stSidebarCollapsedControl"]::before,
        div[data-testid="collapsedControl"]::before,
        button[aria-label="Open sidebar"]::before,
        button[title="Open sidebar"]::before {
            content: "☰" !important;
            pointer-events: none !important;
            color: #FFFFFF !important;
            font-size: 1.55rem !important;
            font-weight: 950 !important;
            line-height: 1 !important;
            text-shadow: 0 2px 8px rgba(0,0,0,.25) !important;
        }

        [data-testid="collapsedControl"] svg,
        [data-testid="stSidebarCollapsedControl"] svg,
        button[aria-label="Open sidebar"] svg,
        button[title="Open sidebar"] svg {
            opacity: 0 !important;
            width: 0 !important;
            height: 0 !important;
            margin: 0 !important;
        }

        [data-testid="collapsedControl"]:hover,
        [data-testid="stSidebarCollapsedControl"]:hover,
        button[aria-label="Open sidebar"]:hover,
        button[title="Open sidebar"]:hover {
            transform: translateY(-1px) scale(1.02) !important;
            filter: brightness(1.06) saturate(1.05) !important;
            box-shadow:
                0 24px 46px rgba(111,81,61,.38),
                0 0 0 9px rgba(212,160,23,.11),
                inset 0 1px 0 rgba(255,255,255,.42) !important;
        }

        /* Boton de cerrar panel: elegante, sin afectar la accion nativa */
        section[data-testid="stSidebar"] button[aria-label="Close sidebar"],
        section[data-testid="stSidebar"] button[title="Close sidebar"],
        section[data-testid="stSidebar"] button[kind="header"] {
            border-radius: 999px !important;
            background: rgba(255,255,255,.86) !important;
            border: 1px solid rgba(31,107,79,.12) !important;
            box-shadow: 0 10px 18px rgba(16,32,25,.08) !important;
        }


        /* Expander premium para acciones de edición/eliminación */
        [data-testid="stExpander"] {
            border: 1px solid rgba(31,107,79,.14) !important;
            border-radius: 24px !important;
            background: linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,250,232,.82)) !important;
            box-shadow: 0 14px 28px rgba(16,32,25,.07) !important;
            overflow: hidden !important;
            margin-bottom: 1rem !important;
        }
        [data-testid="stExpander"] details {
            border-radius: 24px !important;
        }
        [data-testid="stExpander"] summary {
            min-height: 58px !important;
            padding: .8rem 1rem !important;
            font-weight: 950 !important;
            color: #102019 !important;
            background:
                radial-gradient(circle at 20% 10%, rgba(242,209,107,.25), transparent 30%),
                linear-gradient(135deg, rgba(255,255,255,.98), rgba(245,252,244,.96)) !important;
        }
        [data-testid="stExpander"] summary p,
        [data-testid="stExpander"] summary span,
        [data-testid="stExpander"] summary div {
            color: #102019 !important;
            font-weight: 950 !important;
        }
        [data-testid="stExpander"] summary:hover {
            background:
                radial-gradient(circle at 20% 10%, rgba(242,209,107,.36), transparent 34%),
                linear-gradient(135deg, #FFFFFF, #FFF8DD 55%, #F1FAF4) !important;
        }
        [data-testid="stExpander"] svg {
            color: #14513D !important;
            fill: #14513D !important;
        }


        /* =================================================== */
        /* WHATSAPP PREMIUM PREVIEW */
        /* =================================================== */
        .whatsapp-preview {
            border-radius: 28px;
            padding: 1.05rem;
            margin: .8rem 0 1rem 0;
            background:
                radial-gradient(circle at 14% 12%, rgba(242,209,107,.18), transparent 35%),
                linear-gradient(135deg, #0B1612 0%, #14513D 56%, #1F6B4F 82%, #D4A017 100%);
            border: 1px solid rgba(255,255,255,.16);
            box-shadow: 0 24px 48px rgba(16,32,25,.22);
        }
        .whatsapp-preview,
        .whatsapp-preview * {
            color: #FFFFFF !important;
        }
        .whatsapp-preview-top {
            display:flex;
            justify-content:space-between;
            gap:.85rem;
            align-items:flex-start;
            margin-bottom:.8rem;
        }
        .whatsapp-preview-title {
            font-size: 1.02rem;
            font-weight: 950;
            letter-spacing: -.02em;
            margin-bottom:.25rem;
        }
        .whatsapp-preview-sub {
            color: rgba(255,255,255,.80) !important;
            font-size:.84rem;
            line-height:1.42;
        }
        .whatsapp-preview-chip {
            white-space:nowrap;
            border-radius:999px;
            padding:.35rem .65rem;
            font-size:.75rem;
            font-weight:950;
            color:#102019 !important;
            background:rgba(255,255,255,.90);
            border:1px solid rgba(242,209,107,.34);
        }
        .whatsapp-bubble {
            background: linear-gradient(180deg, #FFFFFF 0%, #F7FFF9 100%);
            border-radius: 22px 22px 22px 8px;
            padding: 1rem 1.05rem;
            border: 1px solid rgba(212,160,23,.32);
            box-shadow:
                0 18px 30px rgba(0,0,0,.12),
                inset 0 1px 0 rgba(255,255,255,.75);
            line-height: 1.48;
            font-size: .94rem;
        }
        .whatsapp-bubble,
        .whatsapp-bubble * {
            color: #102019 !important;
        }
        .wa-brand {
            font-weight: 950;
            color:#14513D !important;
            letter-spacing:-.02em;
            font-size:1.02rem;
            margin-bottom:.1rem;
        }
        .wa-subject {
            color:#6B4B00 !important;
            font-weight:900;
            margin-bottom:.55rem;
        }
        .wa-line {
            height:1px;
            background:linear-gradient(90deg, rgba(31,107,79,.24), rgba(212,160,23,.44), transparent);
            margin:.45rem 0 .65rem 0;
        }
        .wa-body {
            white-space: normal;
            color:#102019 !important;
        }
        .whatsapp-status {
            display:inline-flex;
            align-items:center;
            gap:.38rem;
            border-radius:999px;
            padding:.38rem .7rem;
            font-size:.78rem;
            font-weight:950;
            margin-top:.78rem;
            background:rgba(255,255,255,.14);
            border:1px solid rgba(255,255,255,.20);
            color:#FFFFFF !important;
        }
        .whatsapp-status.ok { background: rgba(34,197,94,.20); border-color: rgba(187,247,208,.35); }
        .whatsapp-status.warn { background: rgba(245,158,11,.22); border-color: rgba(253,230,138,.38); }
        .whatsapp-status.err { background: rgba(239,68,68,.20); border-color: rgba(254,202,202,.36); }

        #MainMenu, footer { visibility:hidden; }

        @media (max-width: 900px) {
            .hero-title { font-size:1.8rem; }
            .hero-strip { grid-template-columns:repeat(2,minmax(0,1fr)); }
            .block-container { padding-left:.8rem; padding-right:.8rem; }
            div[data-testid="stRadio"] div[role="radiogroup"] {
                gap: .55rem !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


aplicar_estilo_bradafin()

# ============================================================
# UTILIDADES
# ============================================================

def money(value):
    try:
        return f"${float(value):,.0f}".replace(",", ".")
    except Exception:
        return "$0"


def pct(value):
    try:
        return f"{float(value) * 100:.1f}%"
    except Exception:
        return "0.0%"


def safe(value):
    return html.escape(str(value or ""))


def clear_cache():
    for name in list(globals().keys()):
        obj = globals().get(name)
        if hasattr(obj, "clear") and name.startswith("obtener_"):
            try:
                obj.clear()
            except Exception:
                pass


def section_header(title, subtitle=""):
    st.markdown(f"<div class='section-title'>{safe(title)}</div>", unsafe_allow_html=True)
    if subtitle:
        st.markdown(f"<div class='section-caption'>{safe(subtitle)}</div>", unsafe_allow_html=True)


def kpi_card(label, value, foot="", variant="green"):
    cls = {"green":"kpi-card kpi-green", "gold":"kpi-card kpi-gold", "red":"kpi-card kpi-red", "blue":"kpi-card kpi-blue"}.get(variant, "kpi-card")
    st.markdown(
        f"""
        <div class="{cls}">
            <div class="kpi-label">{safe(label)}</div>
            <div class="kpi-value">{safe(value)}</div>
            <div class="kpi-foot">{safe(foot)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def aplicar_grafica_premium_oscura(fig, height=340):
    """Aplica fondo oscuro premium a gráficas Plotly para mejorar contraste visual."""
    fig.update_layout(
        height=height,
        paper_bgcolor="#0B1612",
        plot_bgcolor="#0B1612",
        font=dict(color="#F8FFF8"),
        title_font=dict(color="#FFFFFF", size=18),
        margin=dict(l=28, r=24, t=58, b=52),
        xaxis=dict(
            color="#EAF8F1",
            title_font=dict(color="#EAF8F1"),
            tickfont=dict(color="#EAF8F1"),
            gridcolor="rgba(255,255,255,.10)",
            zerolinecolor="rgba(242,209,107,.32)",
            linecolor="rgba(255,255,255,.18)",
        ),
        yaxis=dict(
            color="#EAF8F1",
            title_font=dict(color="#EAF8F1"),
            tickfont=dict(color="#EAF8F1"),
            gridcolor="rgba(255,255,255,.14)",
            zerolinecolor="rgba(242,209,107,.32)",
            linecolor="rgba(255,255,255,.18)",
        ),
        hoverlabel=dict(
            bgcolor="#102019",
            font=dict(color="#FFFFFF"),
            bordercolor="#D4A017",
        ),
    )
    fig.update_xaxes(showgrid=False, automargin=True)
    fig.update_yaxes(showgrid=True, automargin=True)
    return fig


def normalizar_fecha(df, cols=("fecha",)):
    df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)
            try:
                df[col] = df[col].dt.tz_convert(None)
            except Exception:
                try:
                    df[col] = df[col].dt.tz_localize(None)
                except Exception:
                    pass
    return df


def periodo_filtro(df, tipo, fecha_base):
    if df is None or df.empty or "fecha" not in df.columns:
        return pd.DataFrame(columns=df.columns if isinstance(df, pd.DataFrame) else [])
    df = normalizar_fecha(df, ["fecha"])
    base = pd.Timestamp(fecha_base)
    if tipo == "Diario":
        inicio = base.normalize()
        fin = inicio + pd.Timedelta(days=1)
    elif tipo == "Semanal":
        inicio = base.normalize() - pd.Timedelta(days=base.weekday())
        fin = inicio + pd.Timedelta(days=7)
    else:
        inicio = pd.Timestamp(base.year, base.month, 1)
        fin = inicio + pd.offsets.MonthBegin(1)
    return df[(df["fecha"] >= inicio) & (df["fecha"] < fin)].copy()


def periodo_anterior_filtro(df, tipo, fecha_base):
    base = pd.Timestamp(fecha_base)
    if tipo == "Diario":
        return periodo_filtro(df, tipo, base - pd.Timedelta(days=1))
    if tipo == "Semanal":
        return periodo_filtro(df, tipo, base - pd.Timedelta(days=7))
    return periodo_filtro(df, tipo, base - pd.offsets.MonthBegin(1))


def get_user_id_email():
    user = st.session_state.get("user")
    if not user:
        return None, ""
    return getattr(user, "id", None), getattr(user, "email", "")


def sincronizar_sesion_supabase():
    """Restaura la sesion autenticada en el cliente de Supabase para que auth.uid() funcione en RLS."""
    access_token = st.session_state.get("bradafin_access_token")
    refresh_token = st.session_state.get("bradafin_refresh_token")
    if not access_token or not refresh_token:
        return False
    try:
        res = supabase.auth.set_session(access_token, refresh_token)
        session = getattr(res, "session", None)
        user = getattr(res, "user", None)
        if session:
            st.session_state["bradafin_access_token"] = getattr(session, "access_token", access_token)
            st.session_state["bradafin_refresh_token"] = getattr(session, "refresh_token", refresh_token)
        if user:
            st.session_state.user = user
        return True
    except Exception:
        try:
            res = supabase.auth.refresh_session(refresh_token)
            session = getattr(res, "session", None)
            user = getattr(res, "user", None)
            if session:
                st.session_state["bradafin_access_token"] = getattr(session, "access_token", access_token)
                st.session_state["bradafin_refresh_token"] = getattr(session, "refresh_token", refresh_token)
            if user:
                st.session_state.user = user
            return bool(session)
        except Exception:
            return False


def limpiar_sesion_local():
    st.session_state.user = None
    st.session_state.pop("bradafin_access_token", None)
    st.session_state.pop("bradafin_refresh_token", None)


def upsert_safe(tabla, payload, on_conflict=None):
    try:
        if on_conflict:
            result = supabase.table(tabla).upsert(payload, on_conflict=on_conflict).execute()
        else:
            result = supabase.table(tabla).upsert(payload).execute()
        clear_cache()
        return True, result
    except Exception as e:
        return False, e


def insert_safe(tabla, payload):
    try:
        result = supabase.table(tabla).insert(payload).execute()
        clear_cache()
        return True, result
    except Exception as e:
        return False, e


def update_safe(tabla, payload, col, val):
    try:
        result = supabase.table(tabla).update(payload).eq(col, val).execute()
        clear_cache()
        return True, result
    except Exception as e:
        return False, e


def delete_safe(tabla, col, val):
    try:
        result = supabase.table(tabla).delete().eq(col, val).execute()
        clear_cache()
        return True, result
    except Exception as e:
        return False, e


def date_widget_value(value, default=None):
    """Convierte valores de Supabase/pandas en date para st.date_input."""
    default = default or date.today()
    try:
        if value is None or value == "":
            return default
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return default
        return ts.date()
    except Exception:
        return default



def id_limpio(valor):
    """Normaliza ids que pueden venir como None, NaN o texto vacío."""
    try:
        if valor is None:
            return None
        if pd.isna(valor):
            return None
    except Exception:
        pass
    valor = str(valor).strip()
    if not valor or valor.lower() in ["none", "nan", "nat"]:
        return None
    return valor

def movimiento_stock_delta(tipo, cantidad):
    """Delta real que un movimiento debe causar en inventario."""
    try:
        qty = float(cantidad or 0)
    except Exception:
        qty = 0.0
    if tipo == "Venta":
        return -qty
    if tipo == "Compra inventario":
        return qty
    return 0.0


def actualizar_estado_cuenta_por_saldo(cuenta, nuevo_saldo):
    """Recalcula estado de una cuenta según saldo y vencimiento."""
    try:
        nuevo_saldo = float(nuevo_saldo or 0)
    except Exception:
        nuevo_saldo = 0.0
    if nuevo_saldo <= 0:
        return "pagada"
    try:
        fv = pd.to_datetime(cuenta.get("fecha_vencimiento"), errors="coerce")
        if pd.notna(fv) and fv < pd.Timestamp.today().normalize():
            return "vencida"
    except Exception:
        pass
    return "abonada" if nuevo_saldo < float(cuenta.get("monto_total", nuevo_saldo) or nuevo_saldo) else "pendiente"


def actualizar_movimiento_con_stock(movimiento_actual, nuevo_payload):
    """Actualiza un movimiento y ajusta inventario reversando el movimiento anterior y aplicando el nuevo."""
    mov = movimiento_actual or {}
    mov_id = mov.get("id")
    if not mov_id:
        return False, "Movimiento inválido."

    old_producto_id = id_limpio(mov.get("producto_id"))
    new_producto_id = id_limpio(nuevo_payload.get("producto_id"))
    old_delta = movimiento_stock_delta(mov.get("tipo"), mov.get("cantidad", 0))
    new_delta = movimiento_stock_delta(nuevo_payload.get("tipo"), nuevo_payload.get("cantidad", 0))

    ok, res = update_safe("bradafin_movimientos", nuevo_payload, "id", mov_id)
    if not ok:
        return ok, res

    # Reversa efecto anterior y aplica efecto nuevo. Si no hay producto, no toca inventario.
    if old_producto_id and old_delta:
        actualizar_stock_producto(old_producto_id, -old_delta)
    if new_producto_id and new_delta:
        actualizar_stock_producto(new_producto_id, new_delta)
    clear_cache()
    return True, res


def eliminar_movimiento_con_stock(movimiento_actual):
    """Elimina movimiento y reversa su efecto en inventario si era venta o compra de inventario."""
    mov = movimiento_actual or {}
    mov_id = mov.get("id")
    if not mov_id:
        return False, "Movimiento inválido."
    ok, res = delete_safe("bradafin_movimientos", "id", mov_id)
    if ok:
        producto_id = id_limpio(mov.get("producto_id"))
        delta = movimiento_stock_delta(mov.get("tipo"), mov.get("cantidad", 0))
        if producto_id and delta:
            actualizar_stock_producto(producto_id, -delta)
        clear_cache()
    return ok, res


def eliminar_cuenta_y_abonos(cuenta_id):
    """Elimina una cuenta junto con sus abonos asociados para evitar bloqueo por relaciones."""
    try:
        supabase.table("bradafin_abonos").delete().eq("cuenta_id", cuenta_id).execute()
    except Exception:
        pass
    return delete_safe("bradafin_cuentas", "id", cuenta_id)

# ============================================================
# SUPABASE DATA
# ============================================================

@st.cache_data(ttl=60, show_spinner=False)
def obtener_negocio(user_id):
    try:
        res = supabase.table("bradafin_negocios").select("*").eq("usuario_id", user_id).order("creado_en", desc=False).limit(1).execute()
        return res.data[0] if res.data else None
    except Exception:
        return None


def crear_negocio(user_id, email, nombre, tipo_negocio, ciudad, telefono, meta_ventas, margen_objetivo):
    payload = {
        "usuario_id": user_id,
        "nombre": nombre.strip(),
        "tipo_negocio": tipo_negocio.strip(),
        "ciudad": ciudad.strip(),
        "telefono": telefono.strip(),
        "correo": email,
        "moneda": CURRENCY,
        "meta_ventas_mensual": float(meta_ventas or 0),
        "margen_objetivo": float(margen_objetivo or 0.30),
        "whatsapp_alertas": True,
        "email_alertas": True,
        "creado_en": datetime.now().isoformat(),
        "actualizado_en": datetime.now().isoformat(),
    }
    ok, res = insert_safe("bradafin_negocios", payload)
    if ok:
        try:
            negocio_id = res.data[0]["id"]
            crear_categorias_default(negocio_id, user_id)
        except Exception:
            pass
    return ok, res


def crear_categorias_default(negocio_id, user_id):
    rows = []
    for tipo, cats in DEFAULT_CATEGORIAS.items():
        for cat in cats:
            rows.append({"negocio_id": negocio_id, "usuario_id": user_id, "tipo": tipo, "nombre": cat, "creado_en": datetime.now().isoformat()})
    if rows:
        insert_safe("bradafin_categorias", rows)

@st.cache_data(ttl=45, show_spinner=False)
def obtener_categorias(negocio_id, tipo=None):
    try:
        q = supabase.table("bradafin_categorias").select("*").eq("negocio_id", negocio_id)
        if tipo:
            q = q.eq("tipo", tipo)
        res = q.order("tipo").order("nombre").execute()
        return pd.DataFrame(res.data or [])
    except Exception:
        return pd.DataFrame(columns=["id", "negocio_id", "tipo", "nombre"])


def lista_categorias(negocio_id, tipo):
    df = obtener_categorias(negocio_id, tipo)
    cats = [str(x).strip() for x in df.get("nombre", pd.Series(dtype=str)).dropna().tolist() if str(x).strip()]
    return cats or DEFAULT_CATEGORIAS.get(tipo, ["Otros"])

@st.cache_data(ttl=30, show_spinner=False)
def obtener_movimientos(negocio_id):
    cols = ["id","negocio_id","usuario_id","fecha","tipo","categoria","monto","metodo_pago","descripcion","producto_id","cantidad","costo_unitario","precio_unitario","creado_en"]
    try:
        res = supabase.table("bradafin_movimientos").select("*").eq("negocio_id", negocio_id).order("fecha", desc=True).execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = normalizar_fecha(df, ["fecha", "creado_en"])
    for c in ["monto", "cantidad", "costo_unitario", "precio_unitario"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    for c in ["tipo", "categoria", "metodo_pago", "descripcion"]:
        df[c] = df[c].fillna("").astype(str)
    return df

@st.cache_data(ttl=30, show_spinner=False)
def obtener_clientes(negocio_id):
    cols = ["id","negocio_id","usuario_id","nombre","documento","telefono","direccion","observaciones","creado_en","actualizado_en"]
    try:
        res = supabase.table("bradafin_clientes").select("*").eq("negocio_id", negocio_id).order("nombre").execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df

@st.cache_data(ttl=30, show_spinner=False)
def obtener_proveedores(negocio_id):
    cols = ["id","negocio_id","usuario_id","nombre","documento","telefono","direccion","observaciones","creado_en","actualizado_en"]
    try:
        res = supabase.table("bradafin_proveedores").select("*").eq("negocio_id", negocio_id).order("nombre").execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df

@st.cache_data(ttl=30, show_spinner=False)
def obtener_cuentas(negocio_id):
    cols = ["id","negocio_id","usuario_id","tipo","tercero_id","tercero_tipo","tercero_nombre","documento","telefono","concepto","monto_total","saldo_pendiente","fecha","fecha_vencimiento","estado","observaciones","creado_en","actualizado_en"]
    try:
        res = supabase.table("bradafin_cuentas").select("*").eq("negocio_id", negocio_id).order("fecha_vencimiento", desc=False).execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = normalizar_fecha(df, ["fecha", "fecha_vencimiento", "creado_en", "actualizado_en"])
    for c in ["monto_total", "saldo_pendiente"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    for c in ["tipo", "tercero_nombre", "documento", "telefono", "concepto", "estado", "observaciones"]:
        df[c] = df[c].fillna("").astype(str)
    return df

@st.cache_data(ttl=30, show_spinner=False)
def obtener_abonos(negocio_id):
    cols = ["id","negocio_id","usuario_id","cuenta_id","fecha","monto","metodo_pago","nota","creado_en"]
    try:
        res = supabase.table("bradafin_abonos").select("*").eq("negocio_id", negocio_id).order("fecha", desc=True).execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = normalizar_fecha(df, ["fecha", "creado_en"])
    df["monto"] = pd.to_numeric(df["monto"], errors="coerce").fillna(0)
    return df

@st.cache_data(ttl=30, show_spinner=False)
def obtener_productos(negocio_id):
    cols = ["id","negocio_id","usuario_id","codigo","nombre","categoria","costo_unitario","precio_venta","stock","stock_minimo","activo","creado_en","actualizado_en"]
    try:
        res = supabase.table("bradafin_productos").select("*").eq("negocio_id", negocio_id).order("nombre").execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    for c in ["costo_unitario", "precio_venta", "stock", "stock_minimo"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    for c in ["codigo", "nombre", "categoria"]:
        df[c] = df[c].fillna("").astype(str)
    if "activo" in df.columns:
        df["activo"] = df["activo"].fillna(True).astype(bool)
    return df

@st.cache_data(ttl=30, show_spinner=False)
def obtener_cajas(negocio_id):
    cols = ["id","negocio_id","usuario_id","fecha","saldo_inicial","saldo_contado","saldo_esperado","diferencia","estado","nota","creado_en","actualizado_en"]
    try:
        res = supabase.table("bradafin_cajas_diarias").select("*").eq("negocio_id", negocio_id).order("fecha", desc=True).execute()
        df = pd.DataFrame(res.data or [])
    except Exception:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = normalizar_fecha(df, ["fecha", "creado_en", "actualizado_en"])
    for c in ["saldo_inicial", "saldo_contado", "saldo_esperado", "diferencia"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df

# ============================================================
# AUTH Y ONBOARDING
# ============================================================

def render_auth():
    col_left, col_right = st.columns([1.05, .95], gap="large")
    with col_left:
        st.markdown("<div class='login-brand-spacer'></div>", unsafe_allow_html=True)
        render_logo_header(width_full=315, width_icon=92, show_name_fallback=False)
        st.markdown("<div class='login-logo-after'></div>", unsafe_allow_html=True)
        st.markdown(
            f"""
            <div class='hero-card'>
                <div class='hero-badge'>BradaFin · microempresas</div>
                <div class='hero-title'>Controle caja, ventas, cartera e inventario.</div>
                <div class='hero-sub'>{APP_TAGLINE}</div>
                <div class='hero-strip'>
                    <div class='hero-mini'><div class='hero-mini-label'>CAJA</div><div class='hero-mini-value'>Apertura y cierre</div></div>
                    <div class='hero-mini'><div class='hero-mini-label'>CARTERA</div><div class='hero-mini-value'>Cobros y abonos</div></div>
                    <div class='hero-mini'><div class='hero-mini-label'>INVENTARIO</div><div class='hero-mini-value'>Stock y rotación</div></div>
                    <div class='hero-mini'><div class='hero-mini-label'>UTILIDAD</div><div class='hero-mini-value'>Margen real</div></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("<div class='soft-card'><div class='section-title'>Promesa comercial</div><div class='section-caption'>No venda una app. Venda control, claridad, menos desorden y mejores decisiones.</div></div>", unsafe_allow_html=True)
    with col_right:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Acceso", "Primero crea tu cuenta. Si ya estás registrado, entra desde Login.")
        tab_reg, tab_login, tab_reset = st.tabs(["Registro", "Login", "Recuperar"])
        with tab_reg:
            email = st.text_input("Correo", key="reg_email")
            password = st.text_input("Contraseña", type="password", key="reg_password")
            password2 = st.text_input("Confirma contraseña", type="password", key="reg_password2")
            if st.button("Crear cuenta BradaFin", type="primary", use_container_width=True):
                if not email or "@" not in email:
                    st.error("Escribe un correo válido.")
                elif len(password or "") < 6:
                    st.error("La contraseña debe tener al menos 6 caracteres.")
                elif password != password2:
                    st.error("Las contraseñas no coinciden.")
                else:
                    try:
                        supabase.auth.sign_up({"email": email.strip(), "password": password})
                        st.success("Cuenta creada. Ahora entra desde Login. Si Supabase exige confirmación, revisa tu correo.")
                    except Exception as e:
                        st.error(f"No pude crear la cuenta: {e}")
        with tab_login:
            email = st.text_input("Correo", key="login_email")
            password = st.text_input("Contraseña", type="password", key="login_password")
            if st.button("Entrar", type="primary", use_container_width=True):
                if not email or not password:
                    st.error("Escribe correo y contraseña.")
                else:
                    try:
                        res = supabase.auth.sign_in_with_password({"email": email.strip(), "password": password})
                        st.session_state.user = res.user
                        if getattr(res, "session", None):
                            st.session_state["bradafin_access_token"] = getattr(res.session, "access_token", None)
                            st.session_state["bradafin_refresh_token"] = getattr(res.session, "refresh_token", None)
                            sincronizar_sesion_supabase()
                        st.rerun()
                    except Exception as e:
                        st.error(f"No pude iniciar sesión: {e}")
        with tab_reset:
            email = st.text_input("Correo de recuperación", key="reset_email")
            if st.button("Enviar enlace de recuperación", use_container_width=True):
                try:
                    supabase.auth.reset_password_for_email(email.strip())
                    st.success("Si el correo existe, se enviará el enlace de recuperación.")
                except Exception as e:
                    st.error(f"No pude enviar recuperación: {e}")
        st.markdown("</div>", unsafe_allow_html=True)


def render_onboarding(user_id, email):
    st.markdown(
        """
        <div class='hero-card' style='color:#FFFFFF !important;'>
            <div class='hero-badge' style='color:#FFFFFF !important;'>Configuración inicial</div>
            <div class='hero-title' style='color:#FFFFFF !important;'>Registre su negocio.</div>
            <div class='hero-sub' style='color:rgba(255,255,255,.93) !important;'>BradaFin organiza caja, ventas, cartera, inventario y utilidad desde una base empresarial.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    with st.form("onboarding_negocio"):
        col1, col2 = st.columns(2)
        with col1:
            nombre = st.text_input("Nombre del negocio", placeholder="Ej: Tienda La 20")
            tipo = st.text_input("Tipo de negocio", placeholder="Ej: Tienda, restaurante, ferretería")
            ciudad = st.text_input("Ciudad", placeholder="Ej: Puerto Asís")
        with col2:
            telefono = st.text_input("WhatsApp del negocio", placeholder="573001112233")
            meta = st.number_input("Meta mensual de ventas", min_value=0.0, step=10000.0, value=0.0)
            margen = st.slider("Margen objetivo", 0.05, 0.80, 0.30, 0.01)
        submit = st.form_submit_button("Crear negocio", type="primary", use_container_width=True)
    if submit:
        if not nombre.strip():
            st.error("Escribe el nombre del negocio.")
        else:
            ok, res = crear_negocio(user_id, email, nombre, tipo or "Microempresa", ciudad or "", telefono or "", meta, margen)
            if ok:
                st.success("Negocio creado. Ya puedes empezar a usar BradaFin.")
                st.rerun()
            else:
                st.error(f"No pude crear el negocio. Revisa si corriste el SQL de Supabase. Detalle: {res}")

# ============================================================
# MÉTRICAS
# ============================================================

def calcular_metricas(df_movs, df_cuentas, df_productos, fecha_base=None, periodo="Mensual"):
    fecha_base = fecha_base or date.today()
    df_p = periodo_filtro(df_movs, periodo, fecha_base)
    ventas = float(df_p[df_p["tipo"] == "Venta"]["monto"].sum()) if not df_p.empty else 0.0
    gastos = float(df_p[df_p["tipo"].isin(["Gasto operativo", "Salida de caja"] )]["monto"].sum()) if not df_p.empty else 0.0
    compras_inv = float(df_p[df_p["tipo"] == "Compra inventario"]["monto"].sum()) if not df_p.empty else 0.0
    costo_ventas = float((df_p[df_p["tipo"] == "Venta"]["costo_unitario"] * df_p[df_p["tipo"] == "Venta"]["cantidad"]).sum()) if not df_p.empty else 0.0
    utilidad_bruta = ventas - costo_ventas
    utilidad_estimada = ventas - costo_ventas - gastos
    margen = utilidad_bruta / ventas if ventas > 0 else 0.0
    cxc = 0.0
    cxp = 0.0
    vencidas = 0
    if df_cuentas is not None and not df_cuentas.empty:
        hoy = pd.Timestamp.today().normalize()
        cxc_df = df_cuentas[(df_cuentas["tipo"] == "Por cobrar") & (df_cuentas["estado"] != "pagada")]
        cxp_df = df_cuentas[(df_cuentas["tipo"] == "Por pagar") & (df_cuentas["estado"] != "pagada")]
        cxc = float(cxc_df["saldo_pendiente"].sum())
        cxp = float(cxp_df["saldo_pendiente"].sum())
        vencidas = int(((cxc_df["fecha_vencimiento"].notna()) & (cxc_df["fecha_vencimiento"] < hoy)).sum())
    stock_bajo = 0
    capital_inventario = 0.0
    if df_productos is not None and not df_productos.empty:
        stock_bajo = int((df_productos["stock"] <= df_productos["stock_minimo"]).sum())
        capital_inventario = float((df_productos["stock"] * df_productos["costo_unitario"]).sum())
    return {
        "ventas": ventas,
        "gastos": gastos,
        "compras_inv": compras_inv,
        "costo_ventas": costo_ventas,
        "utilidad_bruta": utilidad_bruta,
        "utilidad_estimada": utilidad_estimada,
        "margen": margen,
        "cxc": cxc,
        "cxp": cxp,
        "vencidas": vencidas,
        "stock_bajo": stock_bajo,
        "capital_inventario": capital_inventario,
        "df_periodo": df_p,
    }


def generar_alertas_negocio(negocio, df_movs, df_cuentas, df_productos):
    alertas = []
    metricas = calcular_metricas(df_movs, df_cuentas, df_productos, date.today(), "Mensual")
    margen_obj = float((negocio or {}).get("margen_objetivo", 0.30) or 0.30)
    if metricas["ventas"] <= 0:
        alertas.append(("gold", "Todavía no hay ventas registradas este mes. Registre ventas para medir utilidad real."))
    if metricas["margen"] > 0 and metricas["margen"] < margen_obj:
        alertas.append(("red", f"El margen bruto está en {pct(metricas['margen'])}, por debajo del objetivo {pct(margen_obj)}."))
    if metricas["vencidas"] > 0:
        alertas.append(("red", f"Hay {metricas['vencidas']} cuenta(s) por cobrar vencidas. Conviene cobrar hoy."))
    if metricas["stock_bajo"] > 0:
        alertas.append(("gold", f"Hay {metricas['stock_bajo']} producto(s) con stock bajo."))
    meta = float((negocio or {}).get("meta_ventas_mensual", 0) or 0)
    if meta > 0 and metricas["ventas"] >= meta * 0.80 and metricas["ventas"] < meta:
        alertas.append(("green", "Está cerca de cumplir la meta mensual de ventas."))
    if metricas["utilidad_estimada"] < 0 and metricas["ventas"] > 0:
        alertas.append(("red", "La utilidad estimada está negativa. Revise costos y gastos."))
    if not alertas:
        alertas.append(("green", "No hay alertas fuertes. Mantenga el registro de caja, ventas e inventario."))
    return alertas[:6]

# ============================================================
# WHATSAPP, CORREO, PDF, EXPORT
# ============================================================

def limpiar_telefono(tel):
    return re.sub(r"\D+", "", str(tel or ""))


def normalizar_whatsapp_destino(tel):
    """Normaliza número para API/wa.me. Si recibe 10 dígitos colombianos, agrega 57."""
    numero = limpiar_telefono(tel)
    if not numero:
        return ""
    if numero.startswith("00"):
        numero = numero[2:]
    if WHATSAPP_COUNTRY_CODE and len(numero) == 10 and not numero.startswith(WHATSAPP_COUNTRY_CODE):
        numero = f"{WHATSAPP_COUNTRY_CODE}{numero}"
    return numero


def whatsapp_link(telefono, mensaje):
    tel = normalizar_whatsapp_destino(telefono)
    if not tel:
        return ""
    mensaje = limpiar_texto_whatsapp(mensaje) if "limpiar_texto_whatsapp" in globals() else str(mensaje or "")
    return f"https://wa.me/{tel}?text={urllib.parse.quote(mensaje)}"


def mailto_link(email, subject, body):
    if not email:
        return ""
    return f"mailto:{urllib.parse.quote(email)}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"


def whatsapp_api_disponible():
    if not BRADAFIN_AUTO_WHATSAPP:
        return False
    if WHATSAPP_PROVIDER == "twilio":
        return bool(TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_WHATSAPP_FROM)
    if WHATSAPP_PROVIDER in ["meta", "cloud", "meta_cloud"]:
        return bool(META_WHATSAPP_TOKEN and META_WHATSAPP_PHONE_NUMBER_ID)
    return False


def whatsapp_estado_config():
    if not BRADAFIN_AUTO_WHATSAPP:
        return "manual", "WhatsApp automático está apagado. Se mostrarán mensajes premium listos para abrir en WhatsApp."
    if WHATSAPP_PROVIDER == "twilio":
        if whatsapp_api_disponible():
            return "ok", "WhatsApp automático activo con Twilio."
        return "warn", "Faltan credenciales de Twilio para envío automático."
    if WHATSAPP_PROVIDER in ["meta", "cloud", "meta_cloud"]:
        if whatsapp_api_disponible():
            return "ok", "WhatsApp automático activo con Meta Cloud API."
        return "warn", "Faltan credenciales de Meta WhatsApp Cloud API."
    return "warn", "Proveedor de WhatsApp no reconocido. Usa 'twilio' o 'meta'."


def enviar_whatsapp_automatico(telefono, mensaje):
    """Envía WhatsApp real con Twilio o Meta Cloud API. Si no está configurado, no rompe la app."""
    mensaje = limpiar_texto_whatsapp(mensaje)
    destino = normalizar_whatsapp_destino(telefono)
    if not destino:
        return False, "No hay número de WhatsApp válido."

    if not whatsapp_api_disponible():
        return False, "WhatsApp automático no está configurado. El mensaje queda listo para envío manual."

    try:
        if WHATSAPP_PROVIDER == "twilio":
            url = f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_ACCOUNT_SID}/Messages.json"
            from_number = TWILIO_WHATSAPP_FROM if str(TWILIO_WHATSAPP_FROM).startswith("whatsapp:") else f"whatsapp:{TWILIO_WHATSAPP_FROM}"
            to_number = f"whatsapp:+{destino}"
            data = urllib.parse.urlencode({
                "From": from_number,
                "To": to_number,
                "Body": mensaje,
            }).encode("utf-8")
            auth = base64.b64encode(f"{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}".encode("utf-8")).decode("utf-8")
            req = urllib.request.Request(
                url,
                data=data,
                headers={
                    "Authorization": f"Basic {auth}",
                    "Content-Type": "application/x-www-form-urlencoded",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=18) as resp:
                body = resp.read().decode("utf-8", errors="ignore")
            return True, "WhatsApp enviado automáticamente por Twilio."

        if WHATSAPP_PROVIDER in ["meta", "cloud", "meta_cloud"]:
            url = f"https://graph.facebook.com/v19.0/{META_WHATSAPP_PHONE_NUMBER_ID}/messages"
            payload = {
                "messaging_product": "whatsapp",
                "to": destino,
                "type": "text",
                "text": {
                    "preview_url": False,
                    "body": mensaje,
                },
            }
            req = urllib.request.Request(
                url,
                data=json.dumps(payload).encode("utf-8"),
                headers={
                    "Authorization": f"Bearer {META_WHATSAPP_TOKEN}",
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=18) as resp:
                body = resp.read().decode("utf-8", errors="ignore")
            return True, "WhatsApp enviado automáticamente por Meta Cloud API."

        return False, "Proveedor de WhatsApp no reconocido."
    except urllib.error.HTTPError as e:
        try:
            detalle = e.read().decode("utf-8", errors="ignore")
        except Exception:
            detalle = str(e)
        return False, f"WhatsApp API respondió error {e.code}: {detalle[:280]}"
    except Exception as e:
        return False, f"No pude enviar WhatsApp automático: {e}"


def limpiar_texto_whatsapp(texto):
    """Normaliza mensajes para WhatsApp: saltos reales, sin emojis ni marcadores raros."""
    texto = str(texto or "")
    texto = texto.replace("\\n", "\n")
    texto = texto.replace("\r\n", "\n").replace("\r", "\n")
    # Quita emojis/símbolos especiales que algunos proveedores muestran como  .
    texto = re.sub(r"[\U00010000-\U0010ffff]", "", texto)
    texto = texto.replace("━━━━━━━━━━━━━━", "--------------------")
    texto = texto.replace("•", "-")
    # Quita formato markdown de WhatsApp para que el mensaje sea más serio y limpio.
    texto = texto.replace("*", "").replace("_", "")
    # Limpia espacios repetidos sin dañar saltos de línea.
    lineas = [re.sub(r"[ \t]+", " ", linea).strip() for linea in texto.split("\n")]
    # Evita demasiados saltos seguidos.
    salida = []
    vacia_anterior = False
    for linea in lineas:
        if not linea:
            if not vacia_anterior:
                salida.append("")
            vacia_anterior = True
        else:
            salida.append(linea)
            vacia_anterior = False
    return "\n".join(salida).strip()


def fecha_corta(valor):
    try:
        if valor is not None and pd.notna(valor):
            return str(pd.to_datetime(valor).date())
    except Exception:
        pass
    return "sin fecha"


def whatsapp_footer(negocio=None):
    nombre_negocio = (negocio or {}).get("nombre", "tu negocio")
    return f"Mensaje enviado por {nombre_negocio} desde BradaFin."


def mensaje_cobro(negocio, cuenta):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    cliente = cuenta.get("tercero_nombre", "") or "cliente"
    concepto = cuenta.get("concepto", "cuenta pendiente") or "cuenta pendiente"
    venc = fecha_corta(cuenta.get("fecha_vencimiento"))
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
Recordatorio de pago

Hola {cliente}.

Tienes un saldo pendiente registrado:

Saldo: {money(cuenta.get('saldo_pendiente', 0))}
Concepto: {concepto}
Vencimiento: {venc}

Puedes abonar o pagar completo.
Gracias por mantener tu cuenta al dia.

{whatsapp_footer(negocio)}
""")


def mensaje_cuenta_creada_cliente(negocio, cuenta):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    cliente = cuenta.get("tercero_nombre", "") or "cliente"
    concepto = cuenta.get("concepto", "cuenta pendiente") or "cuenta pendiente"
    venc = fecha_corta(cuenta.get("fecha_vencimiento"))
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
Cuenta registrada

Hola {cliente}.

Se registró una cuenta por cobrar a tu nombre:

Valor: {money(cuenta.get('monto_total', 0))}
Concepto: {concepto}
Vencimiento: {venc}

Puedes realizar un abono o pagar completo cuando lo tengas disponible.

{whatsapp_footer(negocio)}
""")


def mensaje_abono_cliente(negocio, cuenta, monto_abono, nuevo_saldo=0):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    cliente = cuenta.get("tercero_nombre", "") or "cliente"
    concepto = cuenta.get("concepto", "cuenta pendiente") or "cuenta pendiente"
    estado = "Cuenta pagada" if float(nuevo_saldo or 0) <= 0 else "Cuenta abonada"
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
{estado}

Hola {cliente}.

Recibimos tu abono correctamente:

Abono recibido: {money(monto_abono)}
Saldo pendiente: {money(nuevo_saldo)}
Concepto: {concepto}

Gracias por tu pago.

{whatsapp_footer(negocio)}
""")


def mensaje_movimiento_comerciante(negocio, tipo, categoria, monto, fecha_mov, metodo, descripcion="", producto="", cantidad=0):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    fecha_txt = fecha_mov.isoformat() if hasattr(fecha_mov, "isoformat") else str(fecha_mov)
    lineas_extra = []
    if producto:
        lineas_extra.append(f"Producto: {producto}")
    try:
        if float(cantidad or 0) > 0:
            lineas_extra.append(f"Cantidad: {cantidad:g}")
    except Exception:
        pass
    extra = ("\n" + "\n".join(lineas_extra)) if lineas_extra else ""
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
Movimiento registrado

Tipo: {tipo}
Categoria: {categoria}
Monto: {money(monto)}
Metodo: {metodo}
Fecha: {fecha_txt}{extra}
Detalle: {descripcion or 'Sin descripcion'}

Tu negocio quedó actualizado.
""")


def mensaje_caja_comerciante(negocio, fecha_caja, inicial, contado, esperado, estado):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    diferencia = float(contado or 0) - float(esperado or 0)
    fecha_txt = fecha_caja.isoformat() if hasattr(fecha_caja, "isoformat") else str(fecha_caja)
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
Caja {estado}

Fecha: {fecha_txt}
Apertura: {money(inicial)}
Saldo esperado: {money(esperado)}
Saldo contado: {money(contado)}
Diferencia: {money(diferencia)}

Control de caja actualizado.
""")


def mensaje_cuenta_comerciante(negocio, cuenta, accion="Cuenta creada"):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
{accion}

Tercero: {cuenta.get('tercero_nombre','')}
Tipo: {cuenta.get('tipo','')}
Concepto: {cuenta.get('concepto','')}
Valor: {money(cuenta.get('monto_total', cuenta.get('saldo_pendiente', 0)))}
Saldo: {money(cuenta.get('saldo_pendiente', 0))}

Cartera actualizada en BradaFin.
""")


def mensaje_resumen_alertas(negocio, alertas, metricas):
    nombre_negocio = (negocio or {}).get("nombre", "BradaFin")
    lineas = []
    for _, txt in (alertas or [])[:5]:
        lineas.append(f"- {txt}")
    señales = "\n".join(lineas) if lineas else "- Sin alertas fuertes."
    return limpiar_texto_whatsapp(f"""
BradaFin | {nombre_negocio}
Resumen de alertas

Ventas del mes: {money(metricas.get('ventas',0))}
Utilidad estimada: {money(metricas.get('utilidad_estimada',0))}
Cartera por cobrar: {money(metricas.get('cxc',0))}
Productos con stock bajo: {metricas.get('stock_bajo',0)}

Señales de hoy:
{señales}

Revise BradaFin para tomar accion.
""")



def render_whatsapp_preview(titulo, mensaje, telefono="", estado="listo", detalle=""):
    status_class = "ok" if estado == "enviado" else "err" if estado == "error" else "warn"
    estado_txt = "Enviado automáticamente" if estado == "enviado" else "No enviado automáticamente" if estado == "error" else "Listo para enviar"
    mensaje_limpio = limpiar_texto_whatsapp(mensaje)
    lineas = mensaje_limpio.split("\n")
    encabezado = lineas[0] if lineas else "BradaFin"
    asunto = lineas[1] if len(lineas) > 1 else safe(titulo)
    cuerpo = safe("\n".join(lineas[2:]).strip()).replace("\n", "<br>")
    tel = normalizar_whatsapp_destino(telefono)

    st.markdown(
        f"""
        <div class='whatsapp-preview'>
            <div class='whatsapp-preview-top'>
                <div>
                    <div class='whatsapp-preview-title'>WhatsApp BradaFin</div>
                    <div class='whatsapp-preview-sub'>Destino: {safe(tel or 'sin número')} · {safe(detalle or estado_txt)}</div>
                </div>
                <div class='whatsapp-preview-chip'>{safe(estado_txt)}</div>
            </div>
            <div class='whatsapp-bubble'>
                <div class='wa-brand'>{safe(encabezado)}</div>
                <div class='wa-subject'>{safe(asunto)}</div>
                <div class='wa-line'></div>
                <div class='wa-body'>{cuerpo}</div>
            </div>
            <div class='whatsapp-status {status_class}'>{safe(estado_txt)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    link = whatsapp_link(telefono, mensaje_limpio)
    if link and estado != "enviado":
        st.markdown(f"[Abrir en WhatsApp]({link})", unsafe_allow_html=True)


def cola_whatsapp_push(titulo, mensaje, telefono="", estado="listo", detalle=""):
    cola = st.session_state.get("bradafin_whatsapp_flash_queue", [])
    cola.append({
        "titulo": titulo,
        "mensaje": mensaje,
        "telefono": telefono,
        "estado": estado,
        "detalle": detalle,
    })
    st.session_state["bradafin_whatsapp_flash_queue"] = cola[-4:]


def procesar_notificacion_whatsapp(titulo, telefono, mensaje, forzar_manual=False):
    """Envía si API está activa; si no, deja vista premium y enlace manual."""
    if forzar_manual or not whatsapp_api_disponible():
        _, detalle = whatsapp_estado_config()
        cola_whatsapp_push(titulo, mensaje, telefono, "listo", detalle)
        return False, detalle
    ok, detalle = enviar_whatsapp_automatico(telefono, mensaje)
    cola_whatsapp_push(titulo, mensaje, telefono, "enviado" if ok else "error", detalle)
    return ok, detalle


def mostrar_whatsapp_flash():
    cola = st.session_state.pop("bradafin_whatsapp_flash_queue", [])
    if not cola:
        return
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Notificaciones WhatsApp", "Mensajes generados por la última acción.")
    for item in cola:
        render_whatsapp_preview(
            item.get("titulo", "WhatsApp"),
            item.get("mensaje", ""),
            item.get("telefono", ""),
            item.get("estado", "listo"),
            item.get("detalle", ""),
        )
    st.markdown("</div>", unsafe_allow_html=True)



def exportar_excel(df, filename="bradafin_export.xlsx", negocio=None, periodo=None, fecha_base=None, metricas=None, df_cuentas=None, df_productos=None):
    buffer = io.BytesIO()
    out = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    negocio = negocio or {}
    periodo = periodo or "Reporte"
    fecha_base = fecha_base or date.today()
    metricas = metricas or {}

    def numero(valor, default=0):
        try:
            return float(valor if valor not in (None, "") else default)
        except Exception:
            return float(default)

    def dato_metrica(clave, default=0):
        try:
            return numero(metricas.get(clave, default), default)
        except Exception:
            return float(default)

    def limpiar_texto_excel(valor):
        if valor is None:
            return ""
        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass
        texto = str(valor)
        texto = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", texto)
        return texto.strip()

    def normalizar_df_excel(data):
        data = data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame()
        for col in data.columns:
            try:
                if pd.api.types.is_datetime64_any_dtype(data[col]):
                    serie = pd.to_datetime(data[col], errors="coerce")
                    try:
                        serie = serie.dt.tz_localize(None)
                    except Exception:
                        pass
                    data[col] = serie.dt.strftime("%Y-%m-%d")
            except Exception:
                pass
            if str(data[col].dtype) == "object":
                data[col] = data[col].map(limpiar_texto_excel)
        return data

    def movimientos_usuario(data):
        data = normalizar_df_excel(data)
        columnas = {
            "fecha": "Fecha",
            "tipo": "Tipo",
            "categoria": "Categoría",
            "monto": "Monto",
            "metodo_pago": "Método de pago",
            "descripcion": "Descripción",
            "cantidad": "Cantidad",
            "costo_unitario": "Costo unitario",
            "precio_unitario": "Precio unitario",
            "creado_en": "Registrado",
        }
        cols = [c for c in columnas if c in data.columns]
        if not cols:
            return pd.DataFrame(columns=["Fecha", "Tipo", "Categoría", "Monto", "Método de pago", "Descripción"])
        data = data[cols].rename(columns=columnas)
        for col in ["Monto", "Cantidad", "Costo unitario", "Precio unitario"]:
            if col in data.columns:
                data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0)
        return data

    def cartera_usuario(data):
        data = normalizar_df_excel(data)
        columnas = {
            "tipo": "Tipo",
            "tercero_nombre": "Tercero",
            "documento": "Documento",
            "telefono": "Teléfono",
            "concepto": "Concepto",
            "monto_total": "Monto total",
            "saldo_pendiente": "Saldo pendiente",
            "fecha": "Fecha",
            "fecha_vencimiento": "Vencimiento",
            "estado": "Estado",
            "observaciones": "Observaciones",
        }
        if data.empty:
            return pd.DataFrame(columns=list(columnas.values()))
        cols = [c for c in columnas if c in data.columns]
        data = data[cols].rename(columns=columnas)
        for col in ["Monto total", "Saldo pendiente"]:
            if col in data.columns:
                data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0)
        return data

    def inventario_usuario(productos, movimientos):
        productos = normalizar_df_excel(productos)
        if productos.empty:
            return pd.DataFrame(columns=["Código", "Producto", "Categoría", "Costo unitario", "Precio venta", "Margen $", "Margen %", "Stock", "Stock mínimo", "Rotación periodo", "Activo"])
        prod = productos.copy()
        for col in ["costo_unitario", "precio_venta", "stock", "stock_minimo"]:
            if col in prod.columns:
                prod[col] = pd.to_numeric(prod[col], errors="coerce").fillna(0)
        prod["margen_$"] = prod.get("precio_venta", 0) - prod.get("costo_unitario", 0)
        if "precio_venta" in prod.columns:
            prod["margen_%"] = prod.apply(lambda r: ((numero(r.get("precio_venta")) - numero(r.get("costo_unitario"))) / numero(r.get("precio_venta"), 1)) if numero(r.get("precio_venta")) > 0 else 0, axis=1)
        else:
            prod["margen_%"] = 0
        rotacion = {}
        try:
            movs = movimientos.copy() if isinstance(movimientos, pd.DataFrame) else pd.DataFrame()
            if not movs.empty and {"producto_id", "cantidad", "tipo"}.issubset(set(movs.columns)):
                ventas = movs[movs["tipo"] == "Venta"].copy()
                ventas["cantidad"] = pd.to_numeric(ventas["cantidad"], errors="coerce").fillna(0)
                rotacion = ventas.groupby("producto_id")["cantidad"].sum().to_dict()
        except Exception:
            rotacion = {}
        prod["rotacion_periodo"] = prod["id"].map(rotacion).fillna(0) if "id" in prod.columns else 0
        columnas = {
            "codigo": "Código",
            "nombre": "Producto",
            "categoria": "Categoría",
            "costo_unitario": "Costo unitario",
            "precio_venta": "Precio venta",
            "margen_$": "Margen $",
            "margen_%": "Margen %",
            "stock": "Stock",
            "stock_minimo": "Stock mínimo",
            "rotacion_periodo": "Rotación periodo",
            "activo": "Activo",
        }
        cols = [c for c in columnas if c in prod.columns]
        return prod[cols].rename(columns=columnas)

    try:
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule

        movimientos = movimientos_usuario(out)
        inventario = inventario_usuario(df_productos if isinstance(df_productos, pd.DataFrame) else pd.DataFrame(), out)
        cartera = cartera_usuario(df_cuentas if isinstance(df_cuentas, pd.DataFrame) else pd.DataFrame())

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            movimientos.to_excel(writer, index=False, sheet_name="Movimientos")
            inventario.to_excel(writer, index=False, sheet_name="Inventario")
            cartera.to_excel(writer, index=False, sheet_name="Cartera")
            wb = writer.book
            ws = wb.create_sheet("Resumen", 0)

            verde_oscuro = "102019"
            verde = "14513D"
            dorado = "D4A017"
            crema2 = "F6FAF4"
            dorado_suave = "FFF4CF"
            blanco = "FFFFFF"
            rojo_suave = "FFF1E8"
            rojo = "C2410C"
            texto = "102019"
            muted = "52675C"
            linea = "DDE8DF"
            money_fmt = '$ #,##0'
            pct_fmt = '0.0%'
            date_fmt = 'yyyy-mm-dd'
            thin = Side(style="thin", color=linea)
            gold_side = Side(style="thin", color=dorado)
            border_soft = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill_dark = PatternFill("solid", fgColor=verde_oscuro)
            fill_green = PatternFill("solid", fgColor=verde)
            fill_pale = PatternFill("solid", fgColor=crema2)
            fill_gold_pale = PatternFill("solid", fgColor=dorado_suave)
            fill_red_pale = PatternFill("solid", fgColor=rojo_suave)
            fill_white = PatternFill("solid", fgColor=blanco)

            for sheet in wb.worksheets:
                sheet.sheet_view.showGridLines = False

            ws.sheet_properties.tabColor = verde
            for col, width in enumerate([15, 17, 15, 17, 15, 17, 15, 17], start=1):
                ws.column_dimensions[get_column_letter(col)].width = width
            for row in range(1, 36):
                ws.row_dimensions[row].height = 21

            ws.merge_cells("A1:H1")
            ws["A1"] = "BradaFin · Reporte empresarial"
            ws["A1"].font = Font(color=blanco, bold=True, size=17)
            ws["A1"].fill = fill_dark
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 31

            ws.merge_cells("A3:B4")
            ws["A3"] = "BradaFin"
            ws["A3"].font = Font(color=verde, bold=True, size=24)
            ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["A3"].fill = fill_white

            ws.merge_cells("C3:H3")
            ws["C3"] = "Excel premium de control empresarial"
            ws["C3"].font = Font(color=texto, bold=True, size=18)
            ws["C3"].alignment = Alignment(horizontal="right", vertical="center")
            ws.merge_cells("C4:H4")
            nombre_negocio = limpiar_texto_excel(negocio.get("nombre", "Negocio") if isinstance(negocio, dict) else "Negocio")
            ws["C4"] = f"{nombre_negocio} · {periodo} · {pd.Timestamp(fecha_base).strftime('%Y-%m-%d')}"
            ws["C4"].font = Font(color=muted, size=10)
            ws["C4"].alignment = Alignment(horizontal="right", vertical="center")

            ventas = dato_metrica("ventas")
            gastos = dato_metrica("gastos")
            utilidad = dato_metrica("utilidad_estimada")
            margen = dato_metrica("margen")
            cxc = dato_metrica("cxc")
            cxp = dato_metrica("cxp")
            capital_inv = dato_metrica("capital_inventario")
            stock_bajo = dato_metrica("stock_bajo")
            costo_ventas = dato_metrica("costo_ventas")

            ws.merge_cells("A6:H6")
            ws["A6"] = "Resumen ejecutivo"
            ws["A6"].font = Font(color=verde, bold=True, size=13)
            ws["A6"].border = Border(bottom=gold_side)

            def pintar_kpi(r1, c1, r2, c2, titulo, valor, tono="green", formato=money_fmt):
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r1, end_column=c2)
                ws.merge_cells(start_row=r1 + 1, start_column=c1, end_row=r2, end_column=c2)
                celda_titulo = ws.cell(r1, c1)
                celda_valor = ws.cell(r1 + 1, c1)
                celda_titulo.value = titulo.upper()
                celda_valor.value = valor
                fill = fill_pale if tono == "green" else fill_gold_pale if tono == "gold" else fill_red_pale if tono == "red" else fill_white
                for row in range(r1, r2 + 1):
                    for col in range(c1, c2 + 1):
                        cell = ws.cell(row, col)
                        cell.fill = fill
                        cell.border = border_soft
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                celda_titulo.font = Font(color=muted, bold=True, size=8)
                celda_valor.font = Font(color=verde if tono == "green" else rojo if tono == "red" else texto, bold=True, size=14)
                celda_valor.number_format = formato

            pintar_kpi(8, 1, 10, 2, "Ventas", ventas, "green")
            pintar_kpi(8, 3, 10, 4, "Gastos", gastos, "red" if gastos else "gold")
            pintar_kpi(8, 5, 10, 6, "Utilidad estimada", utilidad, "green" if utilidad >= 0 else "red")
            pintar_kpi(8, 7, 10, 8, "Margen bruto", margen, "green" if margen >= 0 else "red", pct_fmt)
            pintar_kpi(12, 1, 14, 2, "Cartera por cobrar", cxc, "red" if cxc else "green")
            pintar_kpi(12, 3, 14, 4, "Cuentas por pagar", cxp, "red" if cxp else "green")
            pintar_kpi(12, 5, 14, 6, "Inventario capital", capital_inv, "gold")
            pintar_kpi(12, 7, 14, 8, "Stock bajo", stock_bajo, "red" if stock_bajo else "green", "0")

            ws.merge_cells("A16:H16")
            ws["A16"] = "Lectura rápida"
            ws["A16"].fill = fill_green
            ws["A16"].font = Font(color=blanco, bold=True, size=12)
            ws["A16"].alignment = Alignment(horizontal="left", vertical="center")

            lectura = []
            if ventas <= 0:
                lectura.append("Aún no hay ventas registradas en este periodo.")
            elif utilidad >= 0:
                lectura.append("El periodo muestra utilidad estimada positiva.")
            else:
                lectura.append("El periodo muestra utilidad estimada negativa; revise costos, gastos y precios.")
            if ventas > 0:
                margen_obj = numero(negocio.get("margen_objetivo", 0.30) if isinstance(negocio, dict) else 0.30, 0.30)
                if margen >= margen_obj:
                    lectura.append(f"Margen bruto saludable: {margen * 100:.1f}%.")
                else:
                    lectura.append(f"Margen bruto bajo: {margen * 100:.1f}%. Revise costos o precio de venta.")
            if cxc > 0:
                lectura.append(f"Cartera pendiente por {money(cxc)}. Priorice cobros y abonos.")
            if stock_bajo > 0:
                lectura.append(f"Hay {int(stock_bajo)} producto(s) en stock bajo.")
            if not lectura:
                lectura.append("No hay alertas fuertes en este periodo.")

            for idx, item in enumerate(lectura[:5], start=17):
                ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
                cell = ws.cell(idx, 1)
                cell.value = f"• {limpiar_texto_excel(item)}"
                cell.fill = fill_white
                cell.font = Font(color=texto, size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = Border(bottom=thin)

            ws.merge_cells("A24:C24")
            ws["A24"] = "Balance del periodo"
            ws["A24"].font = Font(color=verde, bold=True, size=12)
            ws["A24"].border = Border(bottom=gold_side)
            ws["A25"] = "Concepto"
            ws["B25"] = "Valor"
            ws["C25"] = "Lectura"
            for cell in ws[25]:
                if cell.column <= 3:
                    cell.fill = fill_dark
                    cell.font = Font(color=blanco, bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(bottom=gold_side)
            rows_balance = [
                ("Ventas", ventas, "Ingreso principal del periodo"),
                ("Costo ventas", costo_ventas, "Costo asociado a productos vendidos"),
                ("Gastos", gastos, "Salidas operativas"),
                ("Utilidad", utilidad, "Resultado estimado"),
            ]
            for i, (concepto, valor, lectura_txt) in enumerate(rows_balance, start=26):
                ws.cell(i, 1).value = concepto
                ws.cell(i, 2).value = valor
                ws.cell(i, 2).number_format = money_fmt
                ws.cell(i, 3).value = lectura_txt
                for col in range(1, 4):
                    cell = ws.cell(i, col)
                    cell.fill = fill_pale if i % 2 == 0 else fill_white
                    cell.border = border_soft
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.font = Font(color=texto)
                ws.cell(i, 2).font = Font(color=verde if valor >= 0 else rojo, bold=True)

            ws.merge_cells("A32:H33")
            ws["A32"] = "Nota: este Excel es de control interno. No reemplaza contabilidad, asesoría tributaria ni facturación electrónica."
            ws["A32"].fill = fill_gold_pale
            ws["A32"].font = Font(color=texto, italic=True, size=9)
            ws["A32"].alignment = Alignment(wrap_text=True, vertical="center")
            ws["A32"].border = Border(left=gold_side, right=gold_side, top=gold_side, bottom=gold_side)

            def estilizar_hoja(nombre_hoja, money_cols=None, pct_cols=None, date_cols=None):
                money_cols = set(money_cols or [])
                pct_cols = set(pct_cols or [])
                date_cols = set(date_cols or [])
                sh = wb[nombre_hoja]
                sh.sheet_view.showGridLines = False
                sh.sheet_properties.tabColor = verde
                sh.freeze_panes = "A2"
                max_row = max(sh.max_row, 1)
                max_col = max(sh.max_column, 1)
                for cell in sh[1]:
                    cell.fill = fill_dark
                    cell.font = Font(color=blanco, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(bottom=gold_side)
                headers = [sh.cell(1, col).value for col in range(1, max_col + 1)]
                for row in range(2, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = sh.cell(row, col)
                        header = headers[col - 1]
                        cell.fill = PatternFill("solid", fgColor=blanco if row % 2 else crema2)
                        cell.border = Border(bottom=thin)
                        cell.font = Font(color=texto, size=10)
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        if header in money_cols:
                            cell.number_format = money_fmt
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        elif header in pct_cols:
                            cell.number_format = pct_fmt
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        elif header in date_cols:
                            cell.number_format = date_fmt
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                widths = {
                    "Fecha": 13, "Registrado": 13, "Tipo": 20, "Categoría": 18, "Monto": 15,
                    "Método de pago": 17, "Descripción": 42, "Cantidad": 12, "Costo unitario": 15,
                    "Precio unitario": 15, "Código": 22, "Producto": 30, "Precio venta": 15,
                    "Margen $": 15, "Margen %": 12, "Stock": 11, "Stock mínimo": 13,
                    "Rotación periodo": 18, "Activo": 10, "Tercero": 28, "Documento": 16,
                    "Teléfono": 16, "Concepto": 30, "Monto total": 16, "Saldo pendiente": 18,
                    "Vencimiento": 14, "Estado": 14, "Observaciones": 34,
                }
                for col in range(1, max_col + 1):
                    header = sh.cell(1, col).value
                    sh.column_dimensions[get_column_letter(col)].width = widths.get(str(header), 14)
                sh.row_dimensions[1].height = 28
                for row in range(2, max_row + 1):
                    sh.row_dimensions[row].height = 23
                if max_row >= 1 and max_col >= 1:
                    sh.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                try:
                    if "Saldo pendiente" in headers and max_row >= 2:
                        col = get_column_letter(headers.index("Saldo pendiente") + 1)
                        sh.conditional_formatting.add(f"{col}2:{col}{max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=fill_red_pale))
                    if "Stock" in headers and max_row >= 2:
                        col = get_column_letter(headers.index("Stock") + 1)
                        sh.conditional_formatting.add(f"{col}2:{col}{max_row}", CellIsRule(operator="lessThanOrEqual", formula=["1"], fill=fill_gold_pale))
                except Exception:
                    pass
                sh.page_setup.orientation = "landscape"
                sh.page_setup.fitToWidth = 1
                sh.page_setup.fitToHeight = 0

            estilizar_hoja("Movimientos", money_cols=["Monto", "Costo unitario", "Precio unitario"], date_cols=["Fecha", "Registrado"])
            estilizar_hoja("Inventario", money_cols=["Costo unitario", "Precio venta", "Margen $"], pct_cols=["Margen %"])
            estilizar_hoja("Cartera", money_cols=["Monto total", "Saldo pendiente"], date_cols=["Fecha", "Vencimiento"])

            ws.freeze_panes = "A6"
            ws.sheet_view.zoomScale = 90
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            wb.active = 0

        buffer.seek(0)
        return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename
    except Exception:
        try:
            raw = normalizar_df_excel(out)
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                raw.to_excel(writer, index=False, sheet_name="datos")
            buffer.seek(0)
            return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename
        except Exception:
            csv = normalizar_df_excel(out).to_csv(index=False).encode("utf-8-sig")
            return csv, "text/csv", filename.replace(".xlsx", ".csv")


def generar_pdf_reporte(negocio, periodo, fecha_base, metricas, df_movs, df_cuentas, df_productos):
    if not REPORTLAB_AVAILABLE:
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=32,
    )

    # Paleta BradaFin premium
    VERDE_OSCURO = colors.HexColor("#102019")
    VERDE = colors.HexColor("#14513D")
    VERDE_MEDIO = colors.HexColor("#1F6B4F")
    DORADO = colors.HexColor("#D4A017")
    DORADO_SUAVE = colors.HexColor("#FFF4CF")
    CREMA = colors.HexColor("#FFFCF3")
    CREMA2 = colors.HexColor("#F6FAF4")
    TEXTO = colors.HexColor("#102019")
    MUTED = colors.HexColor("#52675C")
    LINEA = colors.HexColor("#DDE8DF")
    ROJO = colors.HexColor("#C2410C")

    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "TitleBradaPremium",
        parent=styles["Title"],
        textColor=VERDE_OSCURO,
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        alignment=2,
        spaceAfter=0,
    )
    subtitle = ParagraphStyle(
        "SubtitleBradaPremium",
        parent=styles["BodyText"],
        textColor=MUTED,
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=2,
    )
    h = ParagraphStyle(
        "HBradaPremium",
        parent=styles["Heading2"],
        textColor=VERDE,
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        spaceBefore=8,
        spaceAfter=7,
    )
    p = ParagraphStyle(
        "PBradaPremium",
        parent=styles["BodyText"],
        textColor=TEXTO,
        fontName="Helvetica",
        fontSize=9.2,
        leading=13,
    )
    p_muted = ParagraphStyle(
        "PMutedBradaPremium",
        parent=p,
        textColor=MUTED,
        fontSize=8.3,
        leading=11,
    )
    hero_title = ParagraphStyle(
        "HeroTitleBradaPremium",
        parent=styles["BodyText"],
        textColor=colors.white,
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
    )
    hero_text = ParagraphStyle(
        "HeroTextBradaPremium",
        parent=styles["BodyText"],
        textColor=colors.HexColor("#EAF8F1"),
        fontName="Helvetica",
        fontSize=8.7,
        leading=12,
    )
    card_label = ParagraphStyle(
        "CardLabelBradaPremium",
        parent=styles["BodyText"],
        textColor=MUTED,
        fontName="Helvetica-Bold",
        fontSize=7.4,
        leading=9,
        uppercase=True,
    )
    card_value = ParagraphStyle(
        "CardValueBradaPremium",
        parent=styles["BodyText"],
        textColor=VERDE_OSCURO,
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
    )
    card_value_green = ParagraphStyle(
        "CardValueGreenBradaPremium",
        parent=card_value,
        textColor=VERDE,
    )
    card_value_red = ParagraphStyle(
        "CardValueRedBradaPremium",
        parent=card_value,
        textColor=ROJO,
    )
    table_head = ParagraphStyle(
        "TableHeadBradaPremium",
        parent=styles["BodyText"],
        textColor=colors.white,
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
    )
    table_cell = ParagraphStyle(
        "TableCellBradaPremium",
        parent=styles["BodyText"],
        textColor=TEXTO,
        fontName="Helvetica",
        fontSize=7.3,
        leading=9.2,
    )
    table_money = ParagraphStyle(
        "TableMoneyBradaPremium",
        parent=table_cell,
        fontName="Helvetica-Bold",
        textColor=VERDE_OSCURO,
    )

    def pdf_text(value):
        return safe(value).replace("\n", "<br/>")

    def page_bg(canvas, doc_obj):
        canvas.saveState()
        width, height = A4
        canvas.setFillColor(CREMA)
        canvas.rect(0, 0, width, height, fill=1, stroke=0)
        canvas.setFillColor(CREMA2)
        canvas.rect(0, 0, width, height * 0.78, fill=1, stroke=0)

        # Franja superior sutil premium
        canvas.setFillColor(VERDE_OSCURO)
        canvas.rect(0, height - 18, width, 18, fill=1, stroke=0)
        canvas.setFillColor(DORADO)
        canvas.rect(0, height - 20, width, 2, fill=1, stroke=0)

        # Footer
        canvas.setFillColor(MUTED)
        canvas.setFont("Helvetica", 7.2)
        canvas.drawString(30, 18, "BradaFin - control interno para microempresas")
        canvas.drawRightString(width - 30, 18, f"Pagina {doc_obj.page}")
        canvas.restoreState()

    def make_logo():
        logo_path = None
        try:
            if BRADAFIN_LOGO_FULL_PATH and Path(BRADAFIN_LOGO_FULL_PATH).exists():
                logo_path = BRADAFIN_LOGO_FULL_PATH
            elif BRADAFIN_ICON_PATH and Path(BRADAFIN_ICON_PATH).exists():
                logo_path = BRADAFIN_ICON_PATH
        except Exception:
            logo_path = None

        if logo_path:
            try:
                img = Image(str(logo_path), width=132, height=46, kind="proportional")
                img.hAlign = "LEFT"
                return img
            except Exception:
                pass
        return Paragraph("<b>BradaFin</b>", ParagraphStyle(
            "LogoFallbackBrada",
            parent=styles["Heading1"],
            textColor=VERDE,
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
        ))

    def metric_card(label, value, tone="normal"):
        value_style = card_value_green if tone == "green" else card_value_red if tone == "red" else card_value
        return [
            Paragraph(str(label).upper(), card_label),
            Paragraph(pdf_text(value), value_style),
        ]

    story = []

    negocio_nombre = (negocio or {}).get("nombre", "Negocio")
    fecha_txt = pd.Timestamp(fecha_base).strftime("%Y-%m-%d")
    margen_tono = "green" if float(metricas.get("margen", 0) or 0) >= float((negocio or {}).get("margen_objetivo", 0.30) or 0.30) else "red"
    utilidad_tono = "green" if float(metricas.get("utilidad_estimada", 0) or 0) >= 0 else "red"

    # Header con logo
    header_table = Table(
        [[
            make_logo(),
            [
                Paragraph("Reporte empresarial", title),
                Paragraph(f"{pdf_text(negocio_nombre)}  |  {periodo}  |  {fecha_txt}", subtitle),
            ],
        ]],
        colWidths=[168, 365],
        rowHeights=[54],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.white),
        ("BOX", (0,0), (-1,-1), 0.4, LINEA),
        ("LEFTPADDING", (0,0), (-1,-1), 12),
        ("RIGHTPADDING", (0,0), (-1,-1), 12),
        ("TOPPADDING", (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROUNDEDCORNERS", [14, 14, 14, 14]),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 10))

    # Hero / lectura ejecutiva
    hero_data = [[
        [
            Paragraph("Lectura ejecutiva del periodo", hero_title),
            Paragraph("Resumen para tomar decisiones: ventas, utilidad, cartera, inventario y alertas clave en una sola vista.", hero_text),
        ],
        [
            Paragraph("CONTROL", hero_text),
            Paragraph("Caja - Ventas - Cartera - Inventario", hero_title),
        ],
    ]]
    hero = Table(hero_data, colWidths=[345, 188], rowHeights=[66])
    hero.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), VERDE_OSCURO),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.white),
        ("BOX", (0,0), (-1,-1), 0.1, VERDE_OSCURO),
        ("LINEBEFORE", (1,0), (1,0), 1.0, DORADO),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
        ("RIGHTPADDING", (0,0), (-1,-1), 14),
        ("TOPPADDING", (0,0), (-1,-1), 11),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(hero)
    story.append(Spacer(1, 12))

    # Indicadores en tarjetas
    cards_row_1 = [
        metric_card("Ventas", money(metricas["ventas"]), "green"),
        metric_card("Gastos", money(metricas["gastos"]), "red" if metricas["gastos"] > 0 else "normal"),
        metric_card("Utilidad estimada", money(metricas["utilidad_estimada"]), utilidad_tono),
        metric_card("Margen bruto", pct(metricas["margen"]), margen_tono),
    ]
    cards_row_2 = [
        metric_card("Cartera por cobrar", money(metricas["cxc"]), "red" if metricas["cxc"] > 0 else "normal"),
        metric_card("Cuentas por pagar", money(metricas["cxp"]), "red" if metricas["cxp"] > 0 else "normal"),
        metric_card("Inventario en capital", money(metricas["capital_inventario"]), "normal"),
        metric_card("Stock bajo", str(metricas["stock_bajo"]), "red" if metricas["stock_bajo"] else "green"),
    ]

    card_table = Table(
        [cards_row_1, cards_row_2],
        colWidths=[128, 128, 128, 128],
        rowHeights=[52, 52],
        hAlign="CENTER",
    )
    card_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.white),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#E4EBDD")),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#FFF8DD")),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#F4FBF5")),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(card_table)
    story.append(Spacer(1, 14))

    # Lectura rápida
    story.append(Paragraph("Lectura rápida", h))
    lectura = []
    if metricas["ventas"] <= 0:
        lectura.append("Aún no hay ventas registradas en este periodo.")
    elif metricas["utilidad_estimada"] >= 0:
        lectura.append("El periodo muestra utilidad estimada positiva.")
    else:
        lectura.append("El periodo muestra utilidad estimada negativa; revise costos, gastos y precios.")
    if metricas["margen"] > 0:
        margen_obj = float((negocio or {}).get("margen_objetivo", 0.30) or 0.30)
        if metricas["margen"] >= margen_obj:
            lectura.append(f"El margen bruto está en {pct(metricas['margen'])}, alineado o por encima del objetivo.")
        else:
            lectura.append(f"El margen bruto está en {pct(metricas['margen'])}; conviene revisar costos o precio de venta.")
    if metricas["vencidas"] > 0:
        lectura.append(f"Hay {metricas['vencidas']} cuenta(s) por cobrar vencidas; prioridad comercial: cobrar hoy.")
    if metricas["stock_bajo"] > 0:
        lectura.append(f"Hay {metricas['stock_bajo']} producto(s) con stock bajo; revise reposición.")
    if not lectura:
        lectura.append("No hay alertas fuertes en este periodo.")

    lectura_data = [[Paragraph("Punto clave", table_head), Paragraph("Recomendación", table_head)]]
    for idx, item in enumerate(lectura, start=1):
        lectura_data.append([
            Paragraph(f"{idx:02d}", table_money),
            Paragraph(pdf_text(item), table_cell),
        ])
    lectura_tbl = Table(lectura_data, colWidths=[56, 477])
    lectura_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), VERDE),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("GRID", (0,0), (-1,-1), 0.25, LINEA),
        ("LINEBEFORE", (1,1), (1,-1), 1, DORADO),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(lectura_tbl)
    story.append(Spacer(1, 14))

    # Top productos / inventario si hay datos
    if df_productos is not None and not df_productos.empty:
        try:
            dfp = df_productos.copy()
            dfp["margen_$"] = pd.to_numeric(dfp["precio_venta"], errors="coerce").fillna(0) - pd.to_numeric(dfp["costo_unitario"], errors="coerce").fillna(0)
            ventas = metricas.get("df_periodo", pd.DataFrame())
            rot = ventas[ventas["tipo"] == "Venta"].groupby("producto_id")["cantidad"].sum().to_dict() if isinstance(ventas, pd.DataFrame) and not ventas.empty and "producto_id" in ventas.columns else {}
            dfp["rotacion"] = dfp["id"].map(rot).fillna(0)
            top = dfp.sort_values(["rotacion", "margen_$"], ascending=False).head(5)
            if not top.empty:
                story.append(Paragraph("Inventario destacado", h))
                prod_data = [[
                    Paragraph("Producto", table_head),
                    Paragraph("Stock", table_head),
                    Paragraph("Precio", table_head),
                    Paragraph("Margen $", table_head),
                    Paragraph("Rotacion", table_head),
                ]]
                for _, r in top.iterrows():
                    prod_data.append([
                        Paragraph(pdf_text(str(r.get("nombre", ""))[:34]), table_cell),
                        Paragraph(str(float(r.get("stock", 0) or 0)).rstrip("0").rstrip("."), table_money),
                        Paragraph(money(r.get("precio_venta", 0)), table_money),
                        Paragraph(money(r.get("margen_$", 0)), table_money),
                        Paragraph(str(float(r.get("rotacion", 0) or 0)).rstrip("0").rstrip("."), table_money),
                    ])
                prod_tbl = Table(prod_data, colWidths=[203, 65, 85, 85, 95])
                prod_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), VERDE_OSCURO),
                    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                    ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F6FAF4")]),
                    ("GRID", (0,0), (-1,-1), 0.25, LINEA),
                    ("LINEBELOW", (0,0), (-1,0), 1.1, DORADO),
                    ("LEFTPADDING", (0,0), (-1,-1), 7),
                    ("RIGHTPADDING", (0,0), (-1,-1), 7),
                    ("TOPPADDING", (0,0), (-1,-1), 5),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                ]))
                story.append(prod_tbl)
                story.append(Spacer(1, 12))
        except Exception:
            pass

    # Movimientos del periodo
    story.append(Paragraph("Movimientos del periodo", h))
    df_rep = metricas.get("df_periodo", pd.DataFrame())
    if isinstance(df_rep, pd.DataFrame) and not df_rep.empty:
        df_rep = df_rep.sort_values("fecha", ascending=False).head(28)
    if not isinstance(df_rep, pd.DataFrame) or df_rep.empty:
        story.append(Paragraph("Sin movimientos en este periodo.", p))
    else:
        data = [[
            Paragraph("Fecha", table_head),
            Paragraph("Tipo", table_head),
            Paragraph("Categoria", table_head),
            Paragraph("Monto", table_head),
            Paragraph("Descripcion", table_head),
        ]]
        for r in df_rep.itertuples():
            fecha_val = str(r.fecha.date()) if pd.notna(r.fecha) else ""
            data.append([
                Paragraph(pdf_text(fecha_val), table_cell),
                Paragraph(pdf_text(str(r.tipo)[:22]), table_cell),
                Paragraph(pdf_text(str(r.categoria)[:24]), table_cell),
                Paragraph(pdf_text(money(r.monto)), table_money),
                Paragraph(pdf_text(str(r.descripcion)[:54]), table_cell),
            ])
        tbl = Table(data, colWidths=[58, 86, 92, 78, 219], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), VERDE_OSCURO),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("LINEBELOW", (0,0), (-1,0), 1.2, DORADO),
            ("GRID", (0,0), (-1,-1), 0.25, LINEA),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F6FAF4")]),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(tbl)

    story.append(Spacer(1, 14))
    note = Table(
        [[Paragraph("<b>Nota de control:</b> este reporte es de uso interno. No reemplaza contabilidad, asesoría tributaria ni facturación electrónica.", p_muted)]],
        colWidths=[533],
    )
    note.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#FFF8DD")),
        ("BOX", (0,0), (-1,-1), 0.4, DORADO),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(note)

    doc.build(story, onFirstPage=page_bg, onLaterPages=page_bg)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# OPERACIONES DE NEGOCIO
# ============================================================

def crear_movimiento(negocio_id, user_id, tipo, categoria, monto, fecha_mov, metodo_pago, descripcion, producto_id=None, cantidad=0, costo_unitario=0, precio_unitario=0):
    payload = {
        "negocio_id": negocio_id,
        "usuario_id": user_id,
        "fecha": fecha_mov.isoformat() if hasattr(fecha_mov, "isoformat") else str(fecha_mov),
        "tipo": tipo,
        "categoria": categoria,
        "monto": float(monto or 0),
        "metodo_pago": metodo_pago,
        "descripcion": descripcion.strip(),
        "producto_id": producto_id,
        "cantidad": float(cantidad or 0),
        "costo_unitario": float(costo_unitario or 0),
        "precio_unitario": float(precio_unitario or 0),
        "creado_en": datetime.now().isoformat(),
    }
    return insert_safe("bradafin_movimientos", payload)


def actualizar_stock_producto(producto_id, delta_stock):
    if not producto_id:
        return False, "Sin producto"
    try:
        res = supabase.table("bradafin_productos").select("stock").eq("id", producto_id).limit(1).execute()
        if not res.data:
            return False, "Producto no encontrado"
        stock_actual = float(res.data[0].get("stock", 0) or 0)
        nuevo = stock_actual + float(delta_stock or 0)
        return update_safe("bradafin_productos", {"stock": nuevo, "actualizado_en": datetime.now().isoformat()}, "id", producto_id)
    except Exception as e:
        return False, e


def crear_tercero(tabla, negocio_id, user_id, nombre, documento, telefono, direccion, observaciones):
    payload = {
        "negocio_id": negocio_id,
        "usuario_id": user_id,
        "nombre": nombre.strip(),
        "documento": documento.strip(),
        "telefono": limpiar_telefono(telefono),
        "direccion": direccion.strip(),
        "observaciones": observaciones.strip(),
        "creado_en": datetime.now().isoformat(),
        "actualizado_en": datetime.now().isoformat(),
    }
    return insert_safe(tabla, payload)


def crear_cuenta(negocio_id, user_id, tipo, tercero_id, tercero_tipo, tercero_nombre, documento, telefono, concepto, monto, fecha_cuenta, fecha_venc, observaciones):
    payload = {
        "negocio_id": negocio_id,
        "usuario_id": user_id,
        "tipo": tipo,
        "tercero_id": tercero_id,
        "tercero_tipo": tercero_tipo,
        "tercero_nombre": tercero_nombre,
        "documento": documento,
        "telefono": limpiar_telefono(telefono),
        "concepto": concepto.strip(),
        "monto_total": float(monto or 0),
        "saldo_pendiente": float(monto or 0),
        "fecha": fecha_cuenta.isoformat(),
        "fecha_vencimiento": fecha_venc.isoformat() if fecha_venc else None,
        "estado": "pendiente",
        "observaciones": observaciones.strip(),
        "creado_en": datetime.now().isoformat(),
        "actualizado_en": datetime.now().isoformat(),
    }
    return insert_safe("bradafin_cuentas", payload)


def registrar_abono(negocio_id, user_id, cuenta, monto, fecha_abono, metodo, nota):
    monto = float(monto or 0)
    saldo_actual = float(cuenta.get("saldo_pendiente", 0) or 0)
    if monto <= 0:
        return False, "El abono debe ser mayor a cero."
    if monto > saldo_actual:
        monto = saldo_actual
    payload_abono = {
        "negocio_id": negocio_id,
        "usuario_id": user_id,
        "cuenta_id": cuenta["id"],
        "fecha": fecha_abono.isoformat(),
        "monto": monto,
        "metodo_pago": metodo,
        "nota": nota.strip(),
        "creado_en": datetime.now().isoformat(),
    }
    ok, res = insert_safe("bradafin_abonos", payload_abono)
    if not ok:
        return ok, res
    nuevo_saldo = max(saldo_actual - monto, 0)
    estado = "pagada" if nuevo_saldo <= 0 else "abonada"
    if estado != "pagada" and cuenta.get("fecha_vencimiento") is not None and pd.to_datetime(cuenta.get("fecha_vencimiento")) < pd.Timestamp.today().normalize():
        estado = "vencida"
    update_safe("bradafin_cuentas", {"saldo_pendiente": nuevo_saldo, "estado": estado, "actualizado_en": datetime.now().isoformat()}, "id", cuenta["id"])
    tipo_mov = "Cobro recibido" if cuenta.get("tipo") == "Por cobrar" else "Pago proveedor"
    crear_movimiento(negocio_id, user_id, tipo_mov, cuenta.get("concepto", "Abono"), monto, fecha_abono, metodo, f"Abono: {cuenta.get('tercero_nombre','')}")
    return True, res


def generar_codigo_producto(negocio_id, df_productos):
    n = 1 if df_productos is None or df_productos.empty else len(df_productos.index) + 1
    prefix = str(negocio_id).replace("-", "")[:5].upper()
    return f"BRF-{prefix}-{n:06d}"


def normalizar_codigo_producto(codigo):
    """Limpia lo que llega desde lector/pistola de código de barras o escritura manual."""
    return re.sub(r"[\s\r\n\t]+", "", str(codigo or "").strip())


def buscar_producto_por_codigo(df_productos, codigo):
    codigo_norm = normalizar_codigo_producto(codigo).lower()
    if df_productos is None or df_productos.empty or not codigo_norm:
        return None
    tmp = df_productos.copy()
    tmp["_codigo_norm"] = tmp["codigo"].fillna("").astype(str).map(lambda x: normalizar_codigo_producto(x).lower())
    match = tmp[tmp["_codigo_norm"] == codigo_norm]
    if match.empty:
        return None
    return match.iloc[0].to_dict()


def crear_producto(negocio_id, user_id, codigo, nombre, categoria, costo, precio, stock, stock_min):
    payload = {
        "negocio_id": negocio_id,
        "usuario_id": user_id,
        "codigo": codigo.strip(),
        "nombre": nombre.strip(),
        "categoria": categoria.strip(),
        "costo_unitario": float(costo or 0),
        "precio_venta": float(precio or 0),
        "stock": float(stock or 0),
        "stock_minimo": float(stock_min or 0),
        "activo": True,
        "creado_en": datetime.now().isoformat(),
        "actualizado_en": datetime.now().isoformat(),
    }
    return insert_safe("bradafin_productos", payload)


def generar_barcode_png(codigo):
    if not BARCODE_AVAILABLE or not codigo:
        return None
    try:
        buffer = io.BytesIO()
        Code128 = barcode.get_barcode_class("code128")
        code = Code128(str(codigo), writer=ImageWriter())
        code.write(buffer, options={"write_text": True, "module_height": 12.0, "font_size": 9})
        buffer.seek(0)
        return buffer.getvalue()
    except Exception:
        return None


def riesgo_cliente(cliente_id, df_cuentas, df_abonos):
    if not cliente_id or df_cuentas is None or df_cuentas.empty:
        return 0, "Sin historial", "pill-gray"
    c = df_cuentas[(df_cuentas["tercero_id"] == cliente_id) & (df_cuentas["tipo"] == "Por cobrar")].copy()
    if c.empty:
        return 0, "Sin historial", "pill-gray"
    hoy = pd.Timestamp.today().normalize()
    vencidas = c[(c["estado"] != "pagada") & (c["fecha_vencimiento"].notna()) & (c["fecha_vencimiento"] < hoy)]
    saldo_vencido = float(vencidas["saldo_pendiente"].sum()) if not vencidas.empty else 0.0
    total_hist = len(c.index)
    score = min(100, int(len(vencidas.index) * 25 + (saldo_vencido / max(float(c["monto_total"].sum()), 1)) * 60))
    if score >= 80:
        return score, "No fiar", "pill-red"
    if score >= 60:
        return score, "Riesgo moderado", "pill-red"
    if score >= 30:
        return score, "Pago variable", "pill-gold"
    return score, "Buen pagador", "pill-green"

# ============================================================
# PÁGINAS
# ============================================================

def render_inicio(negocio, user_id, df_movs, df_cuentas, df_productos):
    metricas = calcular_metricas(df_movs, df_cuentas, df_productos, date.today(), "Mensual")
    meta = float(negocio.get("meta_ventas_mensual", 0) or 0)
    avance_meta = metricas["ventas"] / meta if meta > 0 else 0
    st.markdown(
        f"""
        <div class='hero-card'>
            <div class='hero-badge'>BradaFin · panel empresarial</div>
            <div class='hero-title'>{safe(negocio.get('nombre','Negocio'))}</div>
            <div class='hero-sub'>Control de caja, ventas, cartera, inventario y utilidad estimada. La meta no es registrar por registrar: es decidir mejor.</div>
            <div class='hero-strip'>
                <div class='hero-mini'><div class='hero-mini-label'>VENTAS MES</div><div class='hero-mini-value'>{money(metricas['ventas'])}</div></div>
                <div class='hero-mini'><div class='hero-mini-label'>UTILIDAD EST.</div><div class='hero-mini-value'>{money(metricas['utilidad_estimada'])}</div></div>
                <div class='hero-mini'><div class='hero-mini-label'>MARGEN</div><div class='hero-mini-value'>{pct(metricas['margen'])}</div></div>
                <div class='hero-mini'><div class='hero-mini-label'>CARTERA</div><div class='hero-mini-value'>{money(metricas['cxc'])}</div></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    c1,c2,c3,c4 = st.columns(4)
    with c1: kpi_card("Ventas", money(metricas["ventas"]), "Periodo mensual", "green")
    with c2: kpi_card("Gastos", money(metricas["gastos"]), "Gastos operativos", "red")
    with c3: kpi_card("Utilidad estimada", money(metricas["utilidad_estimada"]), f"Margen bruto {pct(metricas['margen'])}", "gold" if metricas["utilidad_estimada"] >= 0 else "red")
    with c4: kpi_card("Inventario", money(metricas["capital_inventario"]), f"Stock bajo: {metricas['stock_bajo']}", "blue")

    col_l, col_r = st.columns([1.05, .95], gap="large")
    with col_l:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Alertas del negocio", "Lo que requiere atención hoy.")
        for color, txt in generar_alertas_negocio(negocio, df_movs, df_cuentas, df_productos):
            st.markdown(f"<span class='pill pill-{color}'>● {safe(txt)}</span><br><br>", unsafe_allow_html=True)
        if meta > 0:
            st.markdown(f"**Meta mensual:** {money(metricas['ventas'])} de {money(meta)}")
            st.progress(float(max(0, min(avance_meta, 1))))
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Últimos movimientos", "Ventas, gastos y caja recientes.")
        if df_movs.empty:
            st.info("Todavía no hay movimientos. Registre una venta o gasto para iniciar.")
        else:
            for _, r in df_movs.sort_values("fecha", ascending=False).head(8).iterrows():
                tipo = r.get("tipo", "")
                cls = "amount-good" if tipo in ["Venta", "Entrada de caja", "Cobro recibido"] else "amount-bad"
                signo = "+" if cls == "amount-good" else "-"
                st.markdown(
                    f"""
                    <div class='movement-row'>
                      <div><div class='movement-title'>{safe(r.get('descripcion') or r.get('categoria') or tipo)}</div>
                      <div class='movement-sub'>{pd.to_datetime(r.get('fecha')).date() if pd.notna(r.get('fecha')) else ''} · {safe(tipo)} · {safe(r.get('categoria'))}</div></div>
                      <div class='{cls}'>{signo}{money(r.get('monto',0))}</div>
                    </div>
                    """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
    with col_r:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Resumen visual", "Rentabilidad, cartera e inventario.")
        df_chart = pd.DataFrame([
            {"métrica":"Ventas", "valor": metricas["ventas"]},
            {"métrica":"Costo ventas", "valor": metricas["costo_ventas"]},
            {"métrica":"Gastos", "valor": metricas["gastos"]},
            {"métrica":"Utilidad", "valor": metricas["utilidad_estimada"]},
        ])
        fig = px.bar(df_chart, x="métrica", y="valor", title="Balance del mes")
        aplicar_grafica_premium_oscura(fig, height=340)
        fig.update_traces(marker_color=["#1F6B4F", "#D4A017", "#C2410C", "#2F8F6B"])
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Acciones rápidas", "Atajos de operación.")
        c1,c2 = st.columns(2)
        with c1:
            if st.button("Registrar venta", type="primary", use_container_width=True): st.session_state.pagina = "Ventas y gastos"; st.rerun()
            if st.button("Crear cliente", use_container_width=True): st.session_state.pagina = "Clientes"; st.rerun()
        with c2:
            if st.button("Abrir caja", use_container_width=True): st.session_state.pagina = "Caja diaria"; st.rerun()
            if st.button("Crear cuenta", use_container_width=True): st.session_state.pagina = "Cuentas"; st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)


def render_caja(negocio, user_id, df_movs, df_cajas):
    negocio_id = negocio["id"]
    st.markdown("<div class='hero-card'><div class='hero-badge'>Caja diaria</div><div class='hero-title'>Apertura, entradas, salidas y cierre.</div><div class='hero-sub'>Muchos negocios pierden control por caja desordenada. Esta vista aterriza el efectivo del día.</div></div>", unsafe_allow_html=True)
    fecha_caja = st.date_input("Fecha de caja", value=date.today())
    df_dia = periodo_filtro(df_movs, "Diario", fecha_caja)
    entradas = float(df_dia[df_dia["tipo"].isin(["Venta", "Entrada de caja", "Cobro recibido"] )]["monto"].sum()) if not df_dia.empty else 0.0
    salidas = float(df_dia[df_dia["tipo"].isin(["Gasto operativo", "Salida de caja", "Compra inventario", "Pago proveedor"] )]["monto"].sum()) if not df_dia.empty else 0.0
    caja_row = None
    if not df_cajas.empty:
        tmp = df_cajas[pd.to_datetime(df_cajas["fecha"]).dt.date == fecha_caja]
        if not tmp.empty:
            caja_row = tmp.iloc[0].to_dict()
    saldo_inicial = float((caja_row or {}).get("saldo_inicial", 0) or 0)
    esperado = saldo_inicial + entradas - salidas
    c1,c2,c3,c4 = st.columns(4)
    with c1: kpi_card("Saldo inicial", money(saldo_inicial), "Apertura", "blue")
    with c2: kpi_card("Entradas", money(entradas), "Ventas y cobros", "green")
    with c3: kpi_card("Salidas", money(salidas), "Gastos y pagos", "red")
    with c4: kpi_card("Saldo esperado", money(esperado), "Caja calculada", "gold")

    col1,col2 = st.columns([.95,1.05], gap="large")
    with col1:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Apertura / cierre", "Registre o corrija el dinero contado.")
        with st.form("caja_form"):
            inicial = st.number_input("Saldo inicial", min_value=0.0, step=1000.0, value=float(saldo_inicial))
            contado = st.number_input("Saldo contado al cierre", min_value=0.0, step=1000.0, value=float((caja_row or {}).get("saldo_contado", 0) or 0))
            estado = st.selectbox("Estado", ["abierta", "cerrada"], index=1 if (caja_row or {}).get("estado") == "cerrada" else 0)
            nota = st.text_area("Nota", value=str((caja_row or {}).get("nota", "") or ""))
            submit = st.form_submit_button("Guardar caja", type="primary", use_container_width=True)
        if submit:
            payload = {"negocio_id": negocio_id, "usuario_id": user_id, "fecha": fecha_caja.isoformat(), "saldo_inicial": inicial, "saldo_contado": contado, "saldo_esperado": inicial + entradas - salidas, "diferencia": contado - (inicial + entradas - salidas), "estado": estado, "nota": nota, "actualizado_en": datetime.now().isoformat()}
            if caja_row:
                ok,res = update_safe("bradafin_cajas_diarias", payload, "id", caja_row["id"])
            else:
                payload["creado_en"] = datetime.now().isoformat()
                ok,res = insert_safe("bradafin_cajas_diarias", payload)
            if ok and bool(negocio.get("whatsapp_alertas", True)):
                msg = mensaje_caja_comerciante(negocio, fecha_caja, inicial, contado, inicial + entradas - salidas, estado)
                procesar_notificacion_whatsapp("Caja actualizada", negocio.get("telefono", ""), msg)
            st.success("Caja guardada.") if ok else st.error(f"No pude guardar caja: {res}")
            if ok: st.rerun()

        if caja_row:
            with st.expander("🔨 Eliminar caja del día", expanded=False):
                st.caption("Elimina solo el registro de cierre/apertura. No borra ventas ni gastos del día.")
                confirmar = st.checkbox("Confirmo eliminar esta caja", key="confirm_del_caja")
                if st.button("Eliminar caja", use_container_width=True, disabled=not confirmar):
                    ok, res = delete_safe("bradafin_cajas_diarias", "id", caja_row["id"])
                    st.success("Caja eliminada.") if ok else st.error(f"No pude eliminar caja: {res}")
                    if ok: st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    with col2:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Movimientos del día", "Base del saldo esperado.")
        if df_dia.empty:
            st.info("No hay movimientos en esta fecha.")
        else:
            show = df_dia[["fecha","tipo","categoria","monto","metodo_pago","descripcion"]].copy()
            show["fecha"] = show["fecha"].dt.strftime("%Y-%m-%d")
            st.dataframe(show, use_container_width=True, hide_index=True)
        st.markdown("</div>", unsafe_allow_html=True)

def render_ventas_gastos(negocio, user_id, df_productos, df_movs):
    negocio_id = negocio["id"]
    st.markdown("<div class='hero-card'><div class='hero-badge'>Ventas y gastos</div><div class='hero-title'>Registre la operación del negocio.</div><div class='hero-sub'>Ventas, gastos, entradas/salidas de caja y compras de inventario.</div></div>", unsafe_allow_html=True)
    tipo = st.radio("Tipo de movimiento", TIPOS_MOVIMIENTO, horizontal=True, key="tipo_movimiento_brf")
    cats = lista_categorias(negocio_id, tipo)
    with st.form("mov_form", clear_on_submit=True):
        c1,c2,c3 = st.columns([.85,.9,1])
        with c1: fecha_mov = st.date_input("Fecha", value=date.today())
        with c2: categoria = st.selectbox("Categoría", cats)
        with c3: metodo = st.selectbox("Método de pago", METODOS_PAGO)
        producto_id = None; cantidad = 0.0; costo = 0.0; precio = 0.0; monto = 0.0
        if tipo in ["Venta", "Compra inventario"] and not df_productos.empty:
            productos_activos = df_productos[df_productos["activo"] == True].copy()
            opciones = {"": "Sin producto específico"}
            for _, r in productos_activos.iterrows():
                opciones[str(r["id"])] = f"{r['codigo']} · {r['nombre']} · stock {r['stock']}"
            producto_id = st.selectbox("Producto", list(opciones.keys()), format_func=lambda x: opciones.get(x, x))
            cantidad = st.number_input("Cantidad", min_value=0.0, step=1.0, value=1.0 if tipo == "Venta" else 0.0)
            if producto_id:
                prod = productos_activos[productos_activos["id"].astype(str) == str(producto_id)].iloc[0]
                costo = float(prod["costo_unitario"] or 0)
                precio = float(prod["precio_venta"] or 0)
                monto_default = precio * cantidad if tipo == "Venta" else costo * cantidad
                monto = st.number_input("Monto total", min_value=0.0, step=1000.0, value=float(monto_default))
            else:
                monto = st.number_input("Monto", min_value=0.0, step=1000.0)
        else:
            monto = st.number_input("Monto", min_value=0.0, step=1000.0)
        descripcion = st.text_input("Descripción", placeholder="Ej: venta mostrador, pago servicios, compra proveedor")
        submit = st.form_submit_button("Guardar movimiento", type="primary", use_container_width=True)
    if submit:
        if monto <= 0:
            st.error("El monto debe ser mayor a cero.")
        else:
            ok,res = crear_movimiento(negocio_id, user_id, tipo, categoria, monto, fecha_mov, metodo, descripcion, producto_id or None, cantidad, costo, precio)
            if ok and producto_id:
                delta = -cantidad if tipo == "Venta" else cantidad if tipo == "Compra inventario" else 0
                if delta:
                    actualizar_stock_producto(producto_id, delta)
            if ok:
                producto_nombre = ""
                try:
                    if producto_id and not df_productos.empty:
                        producto_nombre = df_productos[df_productos["id"].astype(str) == str(producto_id)].iloc[0].get("nombre", "")
                except Exception:
                    producto_nombre = ""
                if bool(negocio.get("whatsapp_alertas", True)):
                    msg = mensaje_movimiento_comerciante(negocio, tipo, categoria, monto, fecha_mov, metodo, descripcion, producto_nombre, cantidad)
                    procesar_notificacion_whatsapp("Movimiento registrado", negocio.get("telefono", ""), msg)
            st.success("Movimiento guardado.") if ok else st.error(f"No pude guardar: {res}")
            if ok: st.rerun()

    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Categorías", "Agrega, edita o elimina categorías propias para este tipo.")
    c1,c2 = st.columns([1, .55])
    with c1: nueva = st.text_input("Nueva categoría", key="new_cat")
    with c2:
        if st.button("Agregar categoría", use_container_width=True):
            if nueva.strip():
                ok,res = insert_safe("bradafin_categorias", {"negocio_id": negocio_id, "usuario_id": user_id, "tipo": tipo, "nombre": nueva.strip(), "creado_en": datetime.now().isoformat()})
                st.success("Categoría agregada.") if ok else st.error(f"No pude agregar: {res}")
                if ok: st.rerun()

    with st.expander("🔨 Editar o eliminar categoría", expanded=False):
        df_cats_tipo = obtener_categorias(negocio_id, tipo)
        if df_cats_tipo.empty:
            st.info("No hay categorías personalizadas para editar.")
        else:
            cat_labels = {str(r["id"]): str(r["nombre"]) for _, r in df_cats_tipo.iterrows()}
            cat_id = st.selectbox("Categoría", list(cat_labels.keys()), format_func=lambda x: cat_labels.get(x, x), key="cat_edit_select")
            cat_actual = df_cats_tipo[df_cats_tipo["id"].astype(str) == str(cat_id)].iloc[0].to_dict()
            nuevo_nombre = st.text_input("Nuevo nombre", value=cat_actual.get("nombre", ""), key="cat_edit_nombre")
            cedit, cdel = st.columns(2)
            with cedit:
                if st.button("Guardar categoría", type="primary", use_container_width=True):
                    if nuevo_nombre.strip():
                        ok,res = update_safe("bradafin_categorias", {"nombre": nuevo_nombre.strip()}, "id", cat_id)
                        st.success("Categoría actualizada.") if ok else st.error(f"No pude actualizar: {res}")
                        if ok: st.rerun()
            with cdel:
                confirmar = st.checkbox("Confirmar", key="cat_delete_confirm")
                if st.button("Eliminar categoría", use_container_width=True, disabled=not confirmar):
                    ok,res = delete_safe("bradafin_categorias", "id", cat_id)
                    st.success("Categoría eliminada.") if ok else st.error(f"No pude eliminar: {res}")
                    if ok: st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    with st.expander("🔨 Editar o eliminar movimientos", expanded=False):
        st.caption("Corrige registros mal digitados. Si cambias producto/cantidad, BradaFin ajusta el inventario.")
        if df_movs.empty:
            st.info("Todavía no hay movimientos para editar.")
        else:
            movs_edit = df_movs.sort_values("fecha", ascending=False).head(200).copy()
            mov_labels = {}
            for _, r in movs_edit.iterrows():
                fecha_txt = pd.to_datetime(r["fecha"]).strftime("%Y-%m-%d") if pd.notna(r["fecha"]) else ""
                desc = str(r.get("descripcion") or r.get("categoria") or "")[:36]
                mov_labels[str(r["id"])] = f"{fecha_txt} · {r.get('tipo','')} · {money(r.get('monto',0))} · {desc}"
            mov_id = st.selectbox("Movimiento a corregir", list(mov_labels.keys()), format_func=lambda x: mov_labels.get(x, x), key="mov_edit_select")
            mov = movs_edit[movs_edit["id"].astype(str) == str(mov_id)].iloc[0].to_dict()
        
            with st.form("mov_edit_form"):
                c1,c2,c3 = st.columns([.8,.9,1])
                with c1:
                    fecha_edit = st.date_input("Fecha", value=date_widget_value(mov.get("fecha")), key="mov_edit_fecha")
                with c2:
                    tipo_edit = st.selectbox("Tipo", TIPOS_MOVIMIENTO, index=TIPOS_MOVIMIENTO.index(mov.get("tipo")) if mov.get("tipo") in TIPOS_MOVIMIENTO else 0, key="mov_edit_tipo")
                cats_edit = lista_categorias(negocio_id, tipo_edit)
                with c3:
                    categoria_edit = st.selectbox("Categoría", cats_edit, index=cats_edit.index(mov.get("categoria")) if mov.get("categoria") in cats_edit else 0, key="mov_edit_cat")
                metodo_edit = st.selectbox("Método de pago", METODOS_PAGO, index=METODOS_PAGO.index(mov.get("metodo_pago")) if mov.get("metodo_pago") in METODOS_PAGO else 0, key="mov_edit_metodo")
        
                producto_edit_id = None
                cantidad_edit = float(mov.get("cantidad", 0) or 0)
                costo_edit = float(mov.get("costo_unitario", 0) or 0)
                precio_edit = float(mov.get("precio_unitario", 0) or 0)
                if tipo_edit in ["Venta", "Compra inventario"] and not df_productos.empty:
                    productos_activos = df_productos[df_productos["activo"] == True].copy()
                    opciones = {"": "Sin producto específico"}
                    for _, r in productos_activos.iterrows():
                        opciones[str(r["id"])] = f"{r['codigo']} · {r['nombre']} · stock {r['stock']}"
                    current_pid = id_limpio(mov.get("producto_id")) or ""
                    producto_edit_id = st.selectbox("Producto", list(opciones.keys()), index=list(opciones.keys()).index(current_pid) if current_pid in opciones else 0, format_func=lambda x: opciones.get(x, x), key="mov_edit_producto")
                    cantidad_edit = st.number_input("Cantidad", min_value=0.0, step=1.0, value=float(mov.get("cantidad", 0) or 0), key="mov_edit_cantidad")
                    if producto_edit_id:
                        prod = productos_activos[productos_activos["id"].astype(str) == str(producto_edit_id)].iloc[0]
                        costo_edit = float(prod["costo_unitario"] or 0)
                        precio_edit = float(prod["precio_venta"] or 0)
                        monto_sugerido = precio_edit * cantidad_edit if tipo_edit == "Venta" else costo_edit * cantidad_edit
                        monto_edit = st.number_input("Monto total", min_value=0.0, step=1000.0, value=float(mov.get("monto", monto_sugerido) or monto_sugerido), key="mov_edit_monto")
                    else:
                        monto_edit = st.number_input("Monto", min_value=0.0, step=1000.0, value=float(mov.get("monto", 0) or 0), key="mov_edit_monto2")
                else:
                    monto_edit = st.number_input("Monto", min_value=0.0, step=1000.0, value=float(mov.get("monto", 0) or 0), key="mov_edit_monto3")
                desc_edit = st.text_input("Descripción", value=str(mov.get("descripcion", "") or ""), key="mov_edit_desc")
                guardar_edit = st.form_submit_button("Guardar corrección", type="primary", use_container_width=True)
        
            if guardar_edit:
                payload = {
                    "fecha": fecha_edit.isoformat(),
                    "tipo": tipo_edit,
                    "categoria": categoria_edit,
                    "monto": float(monto_edit or 0),
                    "metodo_pago": metodo_edit,
                    "descripcion": desc_edit.strip(),
                    "producto_id": producto_edit_id or None,
                    "cantidad": float(cantidad_edit or 0),
                    "costo_unitario": float(costo_edit or 0),
                    "precio_unitario": float(precio_edit or 0),
                }
                ok,res = actualizar_movimiento_con_stock(mov, payload)
                st.success("Movimiento actualizado.") if ok else st.error(f"No pude actualizar: {res}")
                if ok: st.rerun()
        
            st.divider()
            confirmar_mov = st.checkbox("Confirmo eliminar este movimiento", key="mov_delete_confirm")
            if st.button("Eliminar movimiento", use_container_width=True, disabled=not confirmar_mov):
                ok,res = eliminar_movimiento_con_stock(mov)
                st.success("Movimiento eliminado.") if ok else st.error(f"No pude eliminar: {res}")
                if ok: st.rerun()

def render_clientes(negocio, user_id, df_clientes, df_cuentas, df_abonos):
    negocio_id = negocio["id"]
    st.markdown("<div class='hero-card'><div class='hero-badge'>Clientes</div><div class='hero-title'>Clientes por nombre y cédula/NIT.</div><div class='hero-sub'>La calificación de pago es interna del comerciante. No es reporte crediticio.</div></div>", unsafe_allow_html=True)
    col1,col2 = st.columns([.9,1.1], gap="large")
    with col1:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Crear cliente", "Datos mínimos para fiar, cobrar y registrar abonos.")
        with st.form("cliente_form", clear_on_submit=True):
            nombre = st.text_input("Nombre completo")
            documento = st.text_input("Cédula / NIT")
            telefono = st.text_input("WhatsApp", placeholder="573001112233")
            direccion = st.text_input("Dirección")
            obs = st.text_area("Observaciones internas")
            submit = st.form_submit_button("Guardar cliente", type="primary", use_container_width=True)
        if submit:
            if not nombre.strip(): st.error("Escribe el nombre.")
            else:
                ok,res = crear_tercero("bradafin_clientes", negocio_id, user_id, nombre, documento, telefono, direccion, obs)
                st.success("Cliente guardado.") if ok else st.error(f"No pude guardar: {res}")
                if ok: st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

        with st.expander("🔨 Editar o eliminar cliente", expanded=False):
            st.caption("Corrige datos de contacto o elimina registros duplicados.")
            if df_clientes.empty:
                st.info("Aún no hay clientes para editar.")
            else:
                labels = {str(r["id"]): f"{r['nombre']} · {r.get('documento','')}" for _, r in df_clientes.iterrows()}
                cid = st.selectbox("Cliente", list(labels.keys()), format_func=lambda x: labels.get(x, x), key="cliente_edit_select")
                cli = df_clientes[df_clientes["id"].astype(str) == str(cid)].iloc[0].to_dict()
                with st.form("cliente_edit_form"):
                    nombre_e = st.text_input("Nombre completo", value=str(cli.get("nombre", "") or ""))
                    documento_e = st.text_input("Cédula / NIT", value=str(cli.get("documento", "") or ""))
                    telefono_e = st.text_input("WhatsApp", value=str(cli.get("telefono", "") or ""))
                    direccion_e = st.text_input("Dirección", value=str(cli.get("direccion", "") or ""))
                    obs_e = st.text_area("Observaciones internas", value=str(cli.get("observaciones", "") or ""))
                    guardar = st.form_submit_button("Guardar cambios", type="primary", use_container_width=True)
                if guardar:
                    payload = {"nombre": nombre_e.strip(), "documento": documento_e.strip(), "telefono": limpiar_telefono(telefono_e), "direccion": direccion_e.strip(), "observaciones": obs_e.strip(), "actualizado_en": datetime.now().isoformat()}
                    ok,res = update_safe("bradafin_clientes", payload, "id", cid)
                    st.success("Cliente actualizado.") if ok else st.error(f"No pude actualizar: {res}")
                    if ok: st.rerun()
                saldo = 0.0
                if not df_cuentas.empty:
                    saldo = float(df_cuentas[(df_cuentas["tercero_id"] == cid) & (df_cuentas["tipo"] == "Por cobrar") & (df_cuentas["estado"] != "pagada")]["saldo_pendiente"].sum())
                if saldo > 0:
                    st.warning(f"Este cliente tiene saldo pendiente por {money(saldo)}. Si intentas eliminarlo, Supabase puede bloquearlo por historial asociado.")
                confirmar = st.checkbox("Confirmo eliminar este cliente", key="cliente_delete_confirm")
                if st.button("Eliminar cliente", use_container_width=True, disabled=not confirmar):
                    ok,res = delete_safe("bradafin_clientes", "id", cid)
                    st.success("Cliente eliminado.") if ok else st.error(f"No pude eliminar. Puede tener cuentas asociadas. Detalle: {res}")
                    if ok: st.rerun()
    with col2:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Lista de clientes", "Riesgo visible solo para el comerciante.")
        if df_clientes.empty:
            st.info("Aún no hay clientes.")
        else:
            for _, cli in df_clientes.iterrows():
                score, label, pill = riesgo_cliente(cli["id"], df_cuentas, df_abonos)
                saldo = 0.0
                if not df_cuentas.empty:
                    saldo = float(df_cuentas[(df_cuentas["tercero_id"] == cli["id"]) & (df_cuentas["tipo"] == "Por cobrar") & (df_cuentas["estado"] != "pagada")]["saldo_pendiente"].sum())
                st.markdown(f"<div class='movement-row'><div><div class='movement-title'>{safe(cli['nombre'])}</div><div class='movement-sub'>{safe(cli.get('documento'))} · {safe(cli.get('telefono'))} · saldo {money(saldo)}</div></div><div><span class='pill {pill}'>{safe(label)} · {score}</span></div></div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

def render_proveedores(negocio, user_id, df_proveedores):
    negocio_id = negocio["id"]
    st.markdown("<div class='hero-card'><div class='hero-badge'>Proveedores</div><div class='hero-title'>Organice proveedores y cuentas por pagar.</div><div class='hero-sub'>Base para saber a quién debe pagar, cuánto y cuándo.</div></div>", unsafe_allow_html=True)
    col1,col2 = st.columns([.9,1.1], gap="large")
    with col1:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Crear proveedor", "Datos mínimos para cuentas por pagar.")
        with st.form("prov_form", clear_on_submit=True):
            nombre = st.text_input("Nombre / razón social")
            documento = st.text_input("Cédula / NIT")
            telefono = st.text_input("Teléfono")
            direccion = st.text_input("Dirección")
            obs = st.text_area("Observaciones")
            submit = st.form_submit_button("Guardar proveedor", type="primary", use_container_width=True)
        if submit:
            if not nombre.strip(): st.error("Escribe el nombre.")
            else:
                ok,res = crear_tercero("bradafin_proveedores", negocio_id, user_id, nombre, documento, telefono, direccion, obs)
                st.success("Proveedor guardado.") if ok else st.error(f"No pude guardar: {res}")
                if ok: st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

        with st.expander("🔨 Editar o eliminar proveedor", expanded=False):
            st.caption("Corrige datos o elimina proveedores duplicados.")
            if df_proveedores.empty:
                st.info("Aún no hay proveedores para editar.")
            else:
                labels = {str(r["id"]): f"{r['nombre']} · {r.get('documento','')}" for _, r in df_proveedores.iterrows()}
                pid = st.selectbox("Proveedor", list(labels.keys()), format_func=lambda x: labels.get(x, x), key="proveedor_edit_select")
                prov = df_proveedores[df_proveedores["id"].astype(str) == str(pid)].iloc[0].to_dict()
                with st.form("proveedor_edit_form"):
                    nombre_e = st.text_input("Nombre / razón social", value=str(prov.get("nombre", "") or ""))
                    documento_e = st.text_input("Cédula / NIT", value=str(prov.get("documento", "") or ""))
                    telefono_e = st.text_input("Teléfono", value=str(prov.get("telefono", "") or ""))
                    direccion_e = st.text_input("Dirección", value=str(prov.get("direccion", "") or ""))
                    obs_e = st.text_area("Observaciones", value=str(prov.get("observaciones", "") or ""))
                    guardar = st.form_submit_button("Guardar cambios", type="primary", use_container_width=True)
                if guardar:
                    payload = {"nombre": nombre_e.strip(), "documento": documento_e.strip(), "telefono": limpiar_telefono(telefono_e), "direccion": direccion_e.strip(), "observaciones": obs_e.strip(), "actualizado_en": datetime.now().isoformat()}
                    ok,res = update_safe("bradafin_proveedores", payload, "id", pid)
                    st.success("Proveedor actualizado.") if ok else st.error(f"No pude actualizar: {res}")
                    if ok: st.rerun()
                confirmar = st.checkbox("Confirmo eliminar este proveedor", key="proveedor_delete_confirm")
                if st.button("Eliminar proveedor", use_container_width=True, disabled=not confirmar):
                    ok,res = delete_safe("bradafin_proveedores", "id", pid)
                    st.success("Proveedor eliminado.") if ok else st.error(f"No pude eliminar. Puede tener cuentas asociadas. Detalle: {res}")
                    if ok: st.rerun()
    with col2:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Lista de proveedores", "Contactos para compras, crédito y pagos.")
        if df_proveedores.empty:
            st.info("Aún no hay proveedores.")
        else:
            show = df_proveedores[["nombre","documento","telefono","direccion","observaciones"]].copy()
            st.dataframe(show, use_container_width=True, hide_index=True)
        st.markdown("</div>", unsafe_allow_html=True)

def render_cuentas(negocio, user_id, df_clientes, df_proveedores, df_cuentas, df_abonos):
    negocio_id = negocio["id"]
    st.markdown("<div class='hero-card'><div class='hero-badge'>Cuentas y abonos</div><div class='hero-title'>Cobrar, pagar y controlar saldos.</div><div class='hero-sub'>Registre abonos parciales, pagos completos y clientes con riesgo de mora.</div></div>", unsafe_allow_html=True)
    col1,col2 = st.columns([.9,1.1], gap="large")
    with col1:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Crear cuenta", "Por cobrar o por pagar.")
        tipo = st.radio("Tipo", TIPOS_CUENTA, horizontal=True)
        terceros_df = df_clientes if tipo == "Por cobrar" else df_proveedores
        terceros_map = {"": "Crear sin tercero asociado"}
        for _, r in terceros_df.iterrows():
            terceros_map[str(r["id"])] = f"{r['nombre']} · {r.get('documento','')}"
        with st.form("cuenta_form", clear_on_submit=True):
            tercero_id = st.selectbox("Cliente / proveedor", list(terceros_map.keys()), format_func=lambda x: terceros_map.get(x, x))
            nombre_manual = st.text_input("Nombre manual", disabled=bool(tercero_id))
            documento_manual = st.text_input("Documento manual", disabled=bool(tercero_id))
            telefono_manual = st.text_input("WhatsApp manual", disabled=bool(tercero_id))
            concepto = st.text_input("Concepto", placeholder="Ej: venta fiada, factura proveedor")
            monto = st.number_input("Valor", min_value=0.0, step=1000.0)
            c1,c2 = st.columns(2)
            with c1: fecha_cuenta = st.date_input("Fecha", value=date.today())
            with c2: fecha_venc = st.date_input("Vencimiento", value=date.today() + timedelta(days=7))
            obs = st.text_area("Observaciones")
            submit = st.form_submit_button("Guardar cuenta", type="primary", use_container_width=True)
        if submit:
            if tercero_id:
                tercero = terceros_df[terceros_df["id"].astype(str) == str(tercero_id)].iloc[0]
                tercero_nombre = tercero["nombre"]; documento = tercero.get("documento", ""); telefono = tercero.get("telefono", "")
            else:
                tercero_nombre = nombre_manual.strip(); documento = documento_manual.strip(); telefono = telefono_manual.strip()
            if not tercero_nombre or monto <= 0:
                st.error("Escribe tercero y valor mayor a cero.")
            else:
                ok,res = crear_cuenta(negocio_id, user_id, tipo, tercero_id or None, "cliente" if tipo == "Por cobrar" else "proveedor", tercero_nombre, documento, telefono, concepto, monto, fecha_cuenta, fecha_venc, obs)
                if ok:
                    cuenta_notif = {
                        "tipo": tipo,
                        "tercero_nombre": tercero_nombre,
                        "documento": documento,
                        "telefono": telefono,
                        "concepto": concepto,
                        "monto_total": monto,
                        "saldo_pendiente": monto,
                        "fecha_vencimiento": fecha_venc,
                    }
                    if bool(negocio.get("whatsapp_alertas", True)):
                        procesar_notificacion_whatsapp("Cuenta creada para comerciante", negocio.get("telefono", ""), mensaje_cuenta_comerciante(negocio, cuenta_notif, "Cuenta creada"))
                    if tipo == "Por cobrar":
                        procesar_notificacion_whatsapp("Cuenta por cobrar para cliente", telefono, mensaje_cuenta_creada_cliente(negocio, cuenta_notif))
                st.success("Cuenta creada.") if ok else st.error(f"No pude crear cuenta: {res}")
                if ok: st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    with col2:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Cuentas activas", "Seleccione una cuenta para abonar o pagar completo.")
        activas = df_cuentas[df_cuentas["estado"] != "pagada"].copy() if not df_cuentas.empty else pd.DataFrame()
        if activas.empty:
            st.info("No hay cuentas pendientes.")
        else:
            labels = {str(r["id"]): f"{r['tipo']} · {r['tercero_nombre']} · {money(r['saldo_pendiente'])} · vence {pd.to_datetime(r['fecha_vencimiento']).date() if pd.notna(r['fecha_vencimiento']) else 'sin fecha'}" for _, r in activas.iterrows()}
            selected = st.selectbox("Cuenta", list(labels.keys()), format_func=lambda x: labels.get(x, x))
            cuenta = activas[activas["id"].astype(str) == str(selected)].iloc[0].to_dict()
            hoy = pd.Timestamp.today().normalize()
            if pd.notna(cuenta.get("fecha_vencimiento")) and pd.to_datetime(cuenta.get("fecha_vencimiento")) < hoy:
                st.markdown("<span class='pill pill-red'>Cuenta vencida</span>", unsafe_allow_html=True)
            st.markdown(f"**Saldo pendiente:** {money(cuenta.get('saldo_pendiente'))}")
            with st.form("abono_form"):
                monto_abono = st.number_input("Monto del abono / pago", min_value=0.0, step=1000.0, value=float(cuenta.get("saldo_pendiente",0)))
                fecha_abono = st.date_input("Fecha abono", value=date.today())
                metodo = st.selectbox("Método", METODOS_PAGO)
                nota = st.text_input("Nota")
                submit_abono = st.form_submit_button("Registrar abono / pago", type="primary", use_container_width=True)
            if submit_abono:
                saldo_anterior = float(cuenta.get("saldo_pendiente", 0) or 0)
                monto_real = min(float(monto_abono or 0), saldo_anterior)
                ok,res = registrar_abono(negocio_id, user_id, cuenta, monto_abono, fecha_abono, metodo, nota)
                if ok:
                    nuevo_saldo = max(saldo_anterior - monto_real, 0)
                    cuenta_notif = dict(cuenta)
                    cuenta_notif["saldo_pendiente"] = nuevo_saldo
                    if bool(negocio.get("whatsapp_alertas", True)):
                        procesar_notificacion_whatsapp("Abono registrado para comerciante", negocio.get("telefono", ""), mensaje_cuenta_comerciante(negocio, cuenta_notif, f"Abono registrado por {money(monto_real)}"))
                    if cuenta.get("tipo") == "Por cobrar":
                        procesar_notificacion_whatsapp("Recibo de abono para cliente", cuenta.get("telefono", ""), mensaje_abono_cliente(negocio, cuenta, monto_real, nuevo_saldo))
                st.success("Abono registrado.") if ok else st.error(f"No pude registrar abono: {res}")
                if ok: st.rerun()
            if cuenta.get("tipo") == "Por cobrar":
                link = whatsapp_link(cuenta.get("telefono"), mensaje_cobro(negocio, cuenta))
                if link:
                    st.markdown(f"[Enviar recordatorio por WhatsApp]({link})", unsafe_allow_html=True)
                else:
                    st.caption("Esta cuenta no tiene WhatsApp registrado.")
        st.markdown("</div>", unsafe_allow_html=True)

    with st.expander("🔨 Editar o eliminar cuenta", expanded=False):
        st.caption("Corrige valores, fechas, saldo y estado cuando sea necesario.")
        if df_cuentas.empty:
            st.info("No hay cuentas para editar.")
        else:
            labels_all = {str(r["id"]): f"{r['tipo']} · {r['tercero_nombre']} · total {money(r['monto_total'])} · saldo {money(r['saldo_pendiente'])}" for _, r in df_cuentas.iterrows()}
            cuenta_id = st.selectbox("Cuenta a modificar", list(labels_all.keys()), format_func=lambda x: labels_all.get(x, x), key="cuenta_edit_select")
            cuenta_edit = df_cuentas[df_cuentas["id"].astype(str) == str(cuenta_id)].iloc[0].to_dict()
            with st.form("cuenta_edit_form"):
                tipo_e = st.selectbox("Tipo", TIPOS_CUENTA, index=TIPOS_CUENTA.index(cuenta_edit.get("tipo")) if cuenta_edit.get("tipo") in TIPOS_CUENTA else 0)
                tercero_nombre_e = st.text_input("Cliente / proveedor", value=str(cuenta_edit.get("tercero_nombre", "") or ""))
                documento_e = st.text_input("Documento", value=str(cuenta_edit.get("documento", "") or ""))
                telefono_e = st.text_input("WhatsApp", value=str(cuenta_edit.get("telefono", "") or ""))
                concepto_e = st.text_input("Concepto", value=str(cuenta_edit.get("concepto", "") or ""))
                c1,c2 = st.columns(2)
                with c1:
                    monto_total_e = st.number_input("Monto total", min_value=0.0, step=1000.0, value=float(cuenta_edit.get("monto_total", 0) or 0))
                with c2:
                    saldo_e = st.number_input("Saldo pendiente", min_value=0.0, step=1000.0, value=float(cuenta_edit.get("saldo_pendiente", 0) or 0))
                c3,c4 = st.columns(2)
                with c3:
                    fecha_e = st.date_input("Fecha", value=date_widget_value(cuenta_edit.get("fecha")))
                with c4:
                    venc_e = st.date_input("Vencimiento", value=date_widget_value(cuenta_edit.get("fecha_vencimiento"), date.today() + timedelta(days=7)))
                estado_e = st.selectbox("Estado", ESTADOS_CUENTA, index=ESTADOS_CUENTA.index(cuenta_edit.get("estado")) if cuenta_edit.get("estado") in ESTADOS_CUENTA else 0)
                obs_e = st.text_area("Observaciones", value=str(cuenta_edit.get("observaciones", "") or ""))
                guardar = st.form_submit_button("Guardar cambios de cuenta", type="primary", use_container_width=True)
            if guardar:
                payload = {
                    "tipo": tipo_e,
                    "tercero_nombre": tercero_nombre_e.strip(),
                    "documento": documento_e.strip(),
                    "telefono": limpiar_telefono(telefono_e),
                    "concepto": concepto_e.strip(),
                    "monto_total": float(monto_total_e or 0),
                    "saldo_pendiente": float(saldo_e or 0),
                    "fecha": fecha_e.isoformat(),
                    "fecha_vencimiento": venc_e.isoformat() if venc_e else None,
                    "estado": estado_e,
                    "observaciones": obs_e.strip(),
                    "actualizado_en": datetime.now().isoformat(),
                }
                ok,res = update_safe("bradafin_cuentas", payload, "id", cuenta_id)
                st.success("Cuenta actualizada.") if ok else st.error(f"No pude actualizar: {res}")
                if ok: st.rerun()
            st.divider()
            st.caption("Eliminar una cuenta también elimina sus abonos asociados. Los movimientos de caja creados por abonos pueden corregirse desde Ventas y gastos si quedaron registrados.")
            confirmar_cuenta = st.checkbox("Confirmo eliminar esta cuenta y sus abonos", key="cuenta_delete_confirm")
            if st.button("Eliminar cuenta", use_container_width=True, disabled=not confirmar_cuenta):
                ok,res = eliminar_cuenta_y_abonos(cuenta_id)
                st.success("Cuenta eliminada.") if ok else st.error(f"No pude eliminar cuenta: {res}")
                if ok: st.rerun()

    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Historial de abonos", "Fechas y pagos realizados.")
    if df_abonos.empty:
        st.info("Aún no hay abonos registrados.")
    else:
        show = df_abonos[["fecha","monto","metodo_pago","nota","cuenta_id"]].head(50).copy()
        show["fecha"] = show["fecha"].dt.strftime("%Y-%m-%d")
        st.dataframe(show, use_container_width=True, hide_index=True)
    st.markdown("</div>", unsafe_allow_html=True)

def render_inventario(negocio, user_id, df_productos, df_movs):
    negocio_id = negocio["id"]
    st.markdown("<div class='hero-card'><div class='hero-badge'>Inventario</div><div class='hero-title'>Productos, margen, código y rotación.</div><div class='hero-sub'>Controle stock, costo, precio, margen y productos quietos.</div></div>", unsafe_allow_html=True)
    col1,col2 = st.columns([.9,1.1], gap="large")
    with col1:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Crear producto", "Escanee el código si el producto ya lo trae. Si no tiene código, BradaFin genera uno interno.")
        with st.form("producto_form", clear_on_submit=True):
            modo_codigo = st.radio(
                "Código del producto",
                ["Escanear / escribir código existente", "Generar código interno BradaFin"],
                index=0,
                horizontal=False,
                help="La pistola lectora funciona como teclado: haga clic en el campo de código y escanee."
            )
            auto = modo_codigo == "Generar código interno BradaFin"
            if auto:
                st.info("BradaFin generará un código interno para productos que no tienen código de barras comercial.")
                codigo = ""
            else:
                codigo = st.text_input(
                    "Código de barras existente",
                    placeholder="Haz clic aquí y escanea con la pistola, o escribe el código",
                    key="producto_codigo_barras_scan"
                )
                st.caption("Tip: si usas pistola USB/Bluetooth, este campo recibe el código como si fuera teclado. Luego completa nombre, costo, precio y stock.")
            nombre = st.text_input("Nombre del producto")
            categoria = st.text_input("Categoría", value="General")
            c1,c2 = st.columns(2)
            with c1: costo = st.number_input("Costo unitario", min_value=0.0, step=1000.0)
            with c2: precio = st.number_input("Precio venta", min_value=0.0, step=1000.0)
            c3,c4 = st.columns(2)
            with c3: stock = st.number_input("Stock inicial", min_value=0.0, step=1.0)
            with c4: stock_min = st.number_input("Stock mínimo", min_value=0.0, step=1.0)
            submit = st.form_submit_button("Guardar producto", type="primary", use_container_width=True)
        if submit:
            final_codigo = generar_codigo_producto(negocio_id, df_productos) if auto else normalizar_codigo_producto(codigo)
            producto_existente = buscar_producto_por_codigo(df_productos, final_codigo)
            if not nombre.strip():
                st.error("Escribe el nombre del producto.")
            elif not final_codigo:
                st.error("Escanea o escribe el código de barras. Si el producto no tiene código, selecciona 'Generar código interno BradaFin'.")
            elif producto_existente:
                st.error(f"Ya existe un producto con ese código: {producto_existente.get('nombre', 'Producto')} ({producto_existente.get('codigo', final_codigo)}).")
            else:
                ok,res = crear_producto(negocio_id, user_id, final_codigo, nombre, categoria, costo, precio, stock, stock_min)
                st.success(f"Producto guardado con código {final_codigo}.") if ok else st.error(f"No pude guardar: {res}")
                if ok: st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

        with st.expander("🔨 Editar o eliminar producto", expanded=False):
            st.caption("Actualiza precio, costo, stock, código o desactiva productos.")
            if df_productos.empty:
                st.info("Aún no hay productos para editar.")
            else:
                labels = {str(r["id"]): f"{r['codigo']} · {r['nombre']} · stock {r['stock']}" for _, r in df_productos.iterrows()}
                prod_id = st.selectbox("Producto", list(labels.keys()), format_func=lambda x: labels.get(x, x), key="producto_edit_select")
                prod = df_productos[df_productos["id"].astype(str) == str(prod_id)].iloc[0].to_dict()
                with st.form("producto_edit_form"):
                    codigo_e = st.text_input("Código / código de barras", value=str(prod.get("codigo", "") or ""))
                    nombre_e = st.text_input("Nombre del producto", value=str(prod.get("nombre", "") or ""))
                    categoria_e = st.text_input("Categoría", value=str(prod.get("categoria", "") or "General"))
                    c1,c2 = st.columns(2)
                    with c1: costo_e = st.number_input("Costo unitario", min_value=0.0, step=1000.0, value=float(prod.get("costo_unitario", 0) or 0))
                    with c2: precio_e = st.number_input("Precio venta", min_value=0.0, step=1000.0, value=float(prod.get("precio_venta", 0) or 0))
                    c3,c4 = st.columns(2)
                    with c3: stock_e = st.number_input("Stock actual", min_value=0.0, step=1.0, value=float(prod.get("stock", 0) or 0))
                    with c4: stock_min_e = st.number_input("Stock mínimo", min_value=0.0, step=1.0, value=float(prod.get("stock_minimo", 0) or 0))
                    activo_e = st.checkbox("Producto activo", value=bool(prod.get("activo", True)))
                    guardar = st.form_submit_button("Guardar cambios", type="primary", use_container_width=True)
                if guardar:
                    codigo_norm = normalizar_codigo_producto(codigo_e)
                    duplicado = buscar_producto_por_codigo(df_productos[df_productos["id"].astype(str) != str(prod_id)].copy(), codigo_norm)
                    if not nombre_e.strip():
                        st.error("Escribe el nombre del producto.")
                    elif not codigo_norm:
                        st.error("El código no puede quedar vacío.")
                    elif duplicado:
                        st.error(f"Ya existe otro producto con ese código: {duplicado.get('nombre')}.")
                    else:
                        payload = {"codigo": codigo_norm, "nombre": nombre_e.strip(), "categoria": categoria_e.strip(), "costo_unitario": float(costo_e or 0), "precio_venta": float(precio_e or 0), "stock": float(stock_e or 0), "stock_minimo": float(stock_min_e or 0), "activo": bool(activo_e), "actualizado_en": datetime.now().isoformat()}
                        ok,res = update_safe("bradafin_productos", payload, "id", prod_id)
                        st.success("Producto actualizado.") if ok else st.error(f"No pude actualizar: {res}")
                        if ok: st.rerun()
                st.divider()
                st.caption("Si el producto tiene movimientos asociados, Supabase puede bloquear la eliminación. En ese caso BradaFin lo desactiva para no dañar el historial.")
                confirmar = st.checkbox("Confirmo eliminar/desactivar este producto", key="producto_delete_confirm")
                if st.button("Eliminar producto", use_container_width=True, disabled=not confirmar):
                    ok,res = delete_safe("bradafin_productos", "id", prod_id)
                    if ok:
                        st.success("Producto eliminado.")
                        st.rerun()
                    else:
                        ok2,res2 = update_safe("bradafin_productos", {"activo": False, "actualizado_en": datetime.now().isoformat()}, "id", prod_id)
                        st.warning("No se pudo eliminar porque tiene historial. Lo dejé desactivado para que no aparezca al vender.") if ok2 else st.error(f"No pude eliminar ni desactivar: {res}")
                        if ok2: st.rerun()

    with col2:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Rentabilidad y rotación", "Productos que más se mueven y margen.")
        if df_productos.empty:
            st.info("Aún no hay productos.")
        else:
            buscador_codigo = st.text_input(
                "Buscar producto por código / lector",
                placeholder="Escanea aquí para encontrar un producto ya registrado",
                key="buscar_producto_codigo_scan"
            )
            producto_encontrado = buscar_producto_por_codigo(df_productos, buscador_codigo)
            if buscador_codigo:
                if producto_encontrado:
                    st.success(f"Encontrado: {producto_encontrado.get('nombre')} · stock {producto_encontrado.get('stock')} · precio {money(producto_encontrado.get('precio_venta', 0))}")
                else:
                    st.warning("No encontré un producto con ese código. Puedes registrarlo en el formulario de la izquierda.")
            dfp = df_productos.copy()
            dfp["margen_$"] = dfp["precio_venta"] - dfp["costo_unitario"]
            dfp["margen_%"] = dfp.apply(lambda r: (r["precio_venta"] - r["costo_unitario"]) / r["precio_venta"] if r["precio_venta"] else 0, axis=1)
            ventas = df_movs[df_movs["tipo"] == "Venta"].copy() if not df_movs.empty else pd.DataFrame()
            rot = ventas.groupby("producto_id")["cantidad"].sum().to_dict() if not ventas.empty and "producto_id" in ventas.columns else {}
            dfp["rotación"] = dfp["id"].map(rot).fillna(0)
            show = dfp[["codigo","nombre","categoria","costo_unitario","precio_venta","margen_$","margen_%","stock","stock_minimo","rotación","activo"]].copy()
            show["margen_%"] = show["margen_%"].map(lambda x: f"{x*100:.1f}%")
            st.dataframe(show, use_container_width=True, hide_index=True)
            fig = px.bar(dfp.sort_values("rotación", ascending=False).head(10), x="nombre", y="rotación", title="Productos con mayor rotación")
            aplicar_grafica_premium_oscura(fig, height=330)
            fig.update_traces(marker_color="#1F6B4F")
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        st.markdown("</div>", unsafe_allow_html=True)
    if not df_productos.empty:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Código de barras", "Descarga imagen para productos sin código comercial.")
        labels = {str(r["id"]): f"{r['codigo']} · {r['nombre']}" for _, r in df_productos.iterrows()}
        pid = st.selectbox("Producto", list(labels.keys()), format_func=lambda x: labels.get(x, x))
        prod = df_productos[df_productos["id"].astype(str) == str(pid)].iloc[0]
        st.code(prod["codigo"])
        img = generar_barcode_png(prod["codigo"])
        if img:
            st.download_button("Descargar código de barras", data=img, file_name=f"barcode_{prod['codigo']}.png", mime="image/png", use_container_width=True)
        else:
            st.info("Para imagen de código instala python-barcode y pillow. El código interno ya está guardado.")
        st.markdown("</div>", unsafe_allow_html=True)

def render_reportes(negocio, user_id, df_movs, df_cuentas, df_productos):
    st.markdown("<div class='hero-card'><div class='hero-badge'>Reportes</div><div class='hero-title'>Diario, semanal y mensual.</div><div class='hero-sub'>PDF premium, Excel y balance empresarial para tomar decisiones.</div></div>", unsafe_allow_html=True)
    c1,c2 = st.columns([.8,.8])
    with c1: periodo = st.selectbox("Tipo de reporte", ["Diario", "Semanal", "Mensual"], index=2)
    with c2: fecha_base = st.date_input("Fecha base", value=date.today())
    metricas = calcular_metricas(df_movs, df_cuentas, df_productos, fecha_base, periodo)
    metricas_ant = calcular_metricas(periodo_anterior_filtro(df_movs, periodo, fecha_base), df_cuentas, df_productos, fecha_base, periodo)
    c1,c2,c3,c4 = st.columns(4)
    with c1: kpi_card("Ventas", money(metricas["ventas"]), "Periodo seleccionado", "green")
    with c2: kpi_card("Gastos", money(metricas["gastos"]), "Operativos", "red")
    with c3: kpi_card("Utilidad", money(metricas["utilidad_estimada"]), f"Margen {pct(metricas['margen'])}", "gold")
    with c4: kpi_card("Cartera", money(metricas["cxc"]), f"Vencidas {metricas['vencidas']}", "blue")
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Balance del periodo", "Comparación con movimientos registrados.")
    df_chart = pd.DataFrame([
        {"concepto":"Ventas", "valor": metricas["ventas"]},
        {"concepto":"Costo ventas", "valor": metricas["costo_ventas"]},
        {"concepto":"Gastos", "valor": metricas["gastos"]},
        {"concepto":"Utilidad", "valor": metricas["utilidad_estimada"]},
    ])
    fig = px.bar(df_chart, x="concepto", y="valor", title="Rentabilidad monetaria")
    fig.update_traces(marker_color=["#1F6B4F", "#D4A017", "#C2410C", "#2F8F6B"])
    aplicar_grafica_premium_oscura(fig, height=360)
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
    st.markdown("</div>", unsafe_allow_html=True)
    col1,col2 = st.columns(2)
    with col1:
        archivo, mime, fname = exportar_excel(metricas["df_periodo"], f"bradafin_reporte_excel_{periodo.lower()}_{date.today()}.xlsx", negocio=negocio, periodo=periodo, fecha_base=fecha_base, metricas=metricas, df_cuentas=df_cuentas, df_productos=df_productos)
        st.download_button("Descargar Excel premium", data=archivo, file_name=fname, mime=mime, use_container_width=True)
    with col2:
        pdf = generar_pdf_reporte(negocio, periodo, fecha_base, metricas, df_movs, df_cuentas, df_productos)
        if pdf:
            st.download_button("Descargar PDF premium", data=pdf, file_name=f"bradafin_reporte_{periodo.lower()}_{date.today()}.pdf", mime="application/pdf", use_container_width=True)
        else:
            st.info("Para PDF instala reportlab.")


def render_alertas(negocio, user_id, df_movs, df_cuentas, df_productos):
    st.markdown("<div class='hero-card'><div class='hero-badge'>Alertas + WhatsApp</div><div class='hero-title'>Notificaciones automáticas y mensajes premium.</div><div class='hero-sub'>BradaFin puede enviar WhatsApp real con Twilio o Meta Cloud API. Si no hay API configurada, deja el mensaje listo para abrir en WhatsApp.</div></div>", unsafe_allow_html=True)
    alertas = generar_alertas_negocio(negocio, df_movs, df_cuentas, df_productos)
    metricas = calcular_metricas(df_movs, df_cuentas, df_productos, date.today(), "Mensual")

    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Estado de WhatsApp automático", "Configuración actual de envío.")
    estado, detalle = whatsapp_estado_config()
    pill = "green" if estado == "ok" else "gold" if estado in ["warn", "manual"] else "red"
    st.markdown(f"<span class='pill pill-{pill}'>● {safe(detalle)}</span>", unsafe_allow_html=True)
    st.caption("Para envío real necesitas Twilio WhatsApp o Meta WhatsApp Cloud API en Streamlit Secrets. Sin eso, BradaFin conserva el envío manual con mensaje premium.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Alertas internas", "Qué revisar hoy.")
    for color, txt in alertas:
        st.markdown(f"<span class='pill pill-{color}'>● {safe(txt)}</span><br><br>", unsafe_allow_html=True)

    msg_resumen = mensaje_resumen_alertas(negocio, alertas, metricas)
    render_whatsapp_preview("Vista previa para comerciante", msg_resumen, negocio.get("telefono", ""), "listo", "Resumen premium del negocio")
    if st.button("Enviar resumen de alertas al WhatsApp del comerciante", type="primary", use_container_width=True):
        procesar_notificacion_whatsapp("Resumen de alertas", negocio.get("telefono", ""), msg_resumen)
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Cobros por WhatsApp", "Cuentas por cobrar pendientes con mensaje premium.")
    cxc = df_cuentas[(df_cuentas["tipo"] == "Por cobrar") & (df_cuentas["estado"] != "pagada")].copy() if not df_cuentas.empty else pd.DataFrame()
    if cxc.empty:
        st.info("No hay cuentas por cobrar pendientes.")
    else:
        for _, r in cxc.head(12).iterrows():
            cuenta = r.to_dict()
            msg = mensaje_cobro(negocio, cuenta)
            st.markdown(f"**{safe(cuenta.get('tercero_nombre'))}** · {money(cuenta.get('saldo_pendiente'))} · vence {pd.to_datetime(cuenta.get('fecha_vencimiento')).date() if pd.notna(cuenta.get('fecha_vencimiento')) else 'sin fecha'}")
            render_whatsapp_preview("Mensaje de cobro", msg, cuenta.get("telefono", ""), "listo", "Cobro individual")
            cols = st.columns([1, 1])
            with cols[0]:
                if st.button("Enviar automático", key=f"send_cobro_auto_{cuenta.get('id')}", use_container_width=True):
                    procesar_notificacion_whatsapp("Cobro enviado", cuenta.get("telefono", ""), msg)
                    st.rerun()
            with cols[1]:
                link = whatsapp_link(cuenta.get("telefono"), msg)
                if link:
                    st.markdown(f"[Abrir WhatsApp manual]({link})", unsafe_allow_html=True)
                else:
                    st.caption("Sin teléfono válido.")
            st.divider()
    st.markdown("</div>", unsafe_allow_html=True)


def construir_contexto_ia(negocio, df_movs, df_cuentas, df_productos):
    m = calcular_metricas(df_movs, df_cuentas, df_productos, date.today(), "Mensual")
    top_prod = "Sin ventas por producto"
    if not df_movs.empty and not df_productos.empty:
        ventas = df_movs[df_movs["tipo"] == "Venta"].copy()
        if not ventas.empty:
            rot = ventas.groupby("producto_id")["cantidad"].sum().sort_values(ascending=False)
            if not rot.empty and pd.notna(rot.index[0]):
                prod = df_productos[df_productos["id"].astype(str) == str(rot.index[0])]
                if not prod.empty:
                    top_prod = f"{prod.iloc[0]['nombre']} ({rot.iloc[0]} unidades)"
    return f"""
NEGOCIO: {negocio.get('nombre')}
TIPO: {negocio.get('tipo_negocio')}
VENTAS MES: {money(m['ventas'])}
GASTOS MES: {money(m['gastos'])}
COSTO VENTAS: {money(m['costo_ventas'])}
UTILIDAD ESTIMADA: {money(m['utilidad_estimada'])}
MARGEN BRUTO: {pct(m['margen'])}
CARTERA POR COBRAR: {money(m['cxc'])}
CUENTAS POR PAGAR: {money(m['cxp'])}
CUENTAS VENCIDAS: {m['vencidas']}
STOCK BAJO: {m['stock_bajo']}
CAPITAL EN INVENTARIO: {money(m['capital_inventario'])}
PRODUCTO MÁS ROTADO: {top_prod}
""".strip()


def respuesta_ia_local(negocio, df_movs, df_cuentas, df_productos):
    m = calcular_metricas(df_movs, df_cuentas, df_productos, date.today(), "Mensual")
    acciones = []
    if m["vencidas"] > 0:
        acciones.append("cobrar primero las cuentas vencidas")
    if m["stock_bajo"] > 0:
        acciones.append("revisar productos con stock bajo")
    if m["margen"] < float(negocio.get("margen_objetivo", 0.30) or 0.30) and m["ventas"] > 0:
        acciones.append("revisar precios o costos porque el margen está bajo")
    if not acciones:
        acciones.append("mantener registro diario de ventas y gastos")
    return f"Este mes el negocio vendió {money(m['ventas'])}, tuvo gastos por {money(m['gastos'])} y utilidad estimada de {money(m['utilidad_estimada'])}. La acción principal es {acciones[0]}."


def consultar_ia_bradafin(pregunta, contexto, negocio, df_movs, df_cuentas, df_productos):
    if not openai_client:
        return respuesta_ia_local(negocio, df_movs, df_cuentas, df_productos)
    system = """
Eres BradaFin IA, asistente para microempresas. Responde en español colombiano, claro, corto y accionable.
No eres contador ni asesor tributario. No prometas rentabilidades. Enfócate en caja, ventas, gastos, inventario, cartera, margen y decisiones semanales.
Nunca muestres al cliente final etiquetas como mala paga; eso es solo lectura interna del comerciante.
Si no hay datos suficientes, dilo y pide registrar ventas/gastos/productos.
""".strip()
    prompt = f"{contexto}\n\nPregunta: {pregunta}"
    last = None
    for model in GEMINI_MODEL_CANDIDATES:
        try:
            res = openai_client.chat.completions.create(
                model=model,
                messages=[{"role":"system","content":system},{"role":"user","content":prompt}],
                temperature=0.35,
                max_tokens=900,
            )
            txt = res.choices[0].message.content.strip()
            if txt:
                return txt
        except Exception as e:
            last = e
    return respuesta_ia_local(negocio, df_movs, df_cuentas, df_productos) + f"\n\nDetalle técnico IA: {last}"


def render_ia(negocio, user_id, df_movs, df_cuentas, df_productos):
    st.markdown("<div class='hero-card'><div class='hero-badge'>BradaFin IA</div><div class='hero-title'>Pregunte por caja, utilidad, cartera e inventario.</div><div class='hero-sub'>IA empresarial enfocada en decisiones simples para el dueño del negocio.</div></div>", unsafe_allow_html=True)
    if "bradafin_chat" not in st.session_state:
        st.session_state.bradafin_chat = [{"role":"assistant", "content":"Hola. Soy BradaFin IA. Puedo ayudarte a leer si el negocio está ganando, qué cobrar, qué producto revisar y cómo mejorar margen."}]
    col1,col2 = st.columns([1.25,.75], gap="large")
    with col1:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Conversación", "Usa datos reales registrados en BradaFin.")
        for msg in st.session_state.bradafin_chat[-12:]:
            cls = "pill-green" if msg["role"] == "assistant" else "pill-gold"
            st.markdown(f"<div style='padding:.8rem .9rem;border-radius:18px;margin-bottom:.6rem;background:#FFFFFF;border:1px solid rgba(16,32,25,.08);'><strong>{'BradaFin IA' if msg['role']=='assistant' else 'Tú'}</strong><br>{safe(msg['content']).replace(chr(10), '<br>')}</div>", unsafe_allow_html=True)
        with st.form("ia_form"):
            pregunta = st.text_area("Pregunta", placeholder="Ej: ¿Mi negocio está ganando este mes? ¿A quién debo cobrar primero?", height=110)
            send = st.form_submit_button("Preguntar", type="primary", use_container_width=True)
        if send and pregunta.strip():
            st.session_state.bradafin_chat.append({"role":"user", "content":pregunta.strip()})
            with st.spinner("BradaFin está analizando el negocio..."):
                resp = consultar_ia_bradafin(pregunta, construir_contexto_ia(negocio, df_movs, df_cuentas, df_productos), negocio, df_movs, df_cuentas, df_productos)
            st.session_state.bradafin_chat.append({"role":"assistant", "content":resp})
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    with col2:
        st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
        section_header("Preguntas útiles", "Copie una idea.")
        for p in ["¿Mi negocio realmente está ganando?", "¿Qué producto rota más?", "¿Qué producto tiene bajo margen?", "¿A quién debo cobrar primero?", "¿Qué gasto debo revisar esta semana?", "¿Qué precio sugerido debo poner?"]:
            st.markdown(f"<span class='pill pill-gold'>{safe(p)}</span><br><br>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)


def render_perfil(negocio, user_id):
    st.markdown("<div class='hero-card'><div class='hero-badge'>Perfil</div><div class='hero-title'>Configuración del negocio.</div><div class='hero-sub'>Ajuste metas, margen objetivo, alertas y datos de contacto.</div></div>", unsafe_allow_html=True)
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Datos del negocio", "Información visible en reportes y alertas.")
    with st.form("perfil_form"):
        nombre = st.text_input("Nombre", value=negocio.get("nombre", ""))
        tipo = st.text_input("Tipo de negocio", value=negocio.get("tipo_negocio", ""))
        ciudad = st.text_input("Ciudad", value=negocio.get("ciudad", ""))
        telefono = st.text_input("WhatsApp", value=negocio.get("telefono", ""))
        whatsapp_alertas = st.checkbox("Activar notificaciones WhatsApp para movimientos y alertas", value=bool(negocio.get("whatsapp_alertas", True)))
        meta = st.number_input("Meta mensual de ventas", min_value=0.0, step=10000.0, value=float(negocio.get("meta_ventas_mensual", 0) or 0))
        margen = st.slider("Margen objetivo", 0.05, 0.80, float(negocio.get("margen_objetivo", 0.30) or 0.30), 0.01)
        submit = st.form_submit_button("Guardar cambios", type="primary", use_container_width=True)
    if submit:
        payload = {"nombre": nombre, "tipo_negocio": tipo, "ciudad": ciudad, "telefono": telefono, "whatsapp_alertas": whatsapp_alertas, "meta_ventas_mensual": meta, "margen_objetivo": margen, "actualizado_en": datetime.now().isoformat()}
        ok,res = update_safe("bradafin_negocios", payload, "id", negocio["id"])
        st.success("Perfil actualizado.") if ok else st.error(f"No pude actualizar: {res}")
        if ok: st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("WhatsApp automático", "Estado técnico del envío real.")
    estado_w, detalle_w = whatsapp_estado_config()
    pill_w = "green" if estado_w == "ok" else "gold"
    st.markdown(f"<span class='pill pill-{pill_w}'>● {safe(detalle_w)}</span>", unsafe_allow_html=True)
    st.caption("Secrets recomendados: BRADAFIN_AUTO_WHATSAPP, WHATSAPP_PROVIDER, WHATSAPP_COUNTRY_CODE y credenciales de Twilio o Meta.")
    msg_prueba = mensaje_resumen_alertas(negocio, [("green", "Prueba de mensaje premium BradaFin.")], calcular_metricas(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), date.today(), "Mensual"))
    render_whatsapp_preview("Mensaje de prueba", msg_prueba, negocio.get("telefono", ""), "listo", "Así se verá el mensaje en WhatsApp")
    if st.button("Probar notificación al WhatsApp del negocio", use_container_width=True):
        procesar_notificacion_whatsapp("Prueba WhatsApp", negocio.get("telefono", ""), msg_prueba)
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    section_header("Sesión", "Cerrar sesión.")
    if st.button("Cerrar sesión", use_container_width=True):
        try:
            supabase.auth.sign_out()
        except Exception:
            pass
        limpiar_sesion_local()
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)



def render_sidebar_onboarding(email):
    """Mantiene contenido en el sidebar durante la configuracion inicial.
    Sin contenido de sidebar, Streamlit puede ocultar el boton hamburguesa.
    """
    with st.sidebar:
        render_logo_sidebar(width_full=200, width_icon=74, show_name_fallback=not bool(BRADAFIN_LOGO_FULL_PATH))
        st.caption(APP_TAGLINE)
        st.divider()
        st.markdown("**Configuración inicial**")
        st.caption("Crea el negocio para activar el panel completo: caja, ventas, clientes, inventario, reportes y alertas.")
        if email:
            st.caption(email)
        st.divider()
        if st.button("Cerrar sesión", use_container_width=True):
            try:
                supabase.auth.sign_out()
            except Exception:
                pass
            st.session_state.user = None
            st.rerun()


# ============================================================
# MAIN
# ============================================================

def main():
    if "user" not in st.session_state:
        st.session_state.user = None
    if "pagina" not in st.session_state:
        st.session_state.pagina = "Inicio"
    if st.session_state.user is None:
        render_auth()
        st.stop()
    user_id, email = get_user_id_email()
    if not user_id:
        st.error("No pude identificar el usuario activo.")
        st.stop()
    sesion_ok = sincronizar_sesion_supabase()
    if not sesion_ok:
        st.warning("Tu sesión no está sincronizada con Supabase. Cierra sesión y vuelve a entrar para guardar datos correctamente.")
    negocio = obtener_negocio(user_id)
    if not negocio:
        render_sidebar_onboarding(email)
        render_onboarding(user_id, email)
        st.stop()
    negocio_id = negocio["id"]
    df_movs = obtener_movimientos(negocio_id)
    df_clientes = obtener_clientes(negocio_id)
    df_proveedores = obtener_proveedores(negocio_id)
    df_cuentas = obtener_cuentas(negocio_id)
    df_abonos = obtener_abonos(negocio_id)
    df_productos = obtener_productos(negocio_id)
    df_cajas = obtener_cajas(negocio_id)

    with st.sidebar:
        render_logo_sidebar(width_full=200, width_icon=74, show_name_fallback=not bool(BRADAFIN_LOGO_FULL_PATH))
        st.caption(APP_TAGLINE)
        st.markdown(f"**{safe(negocio.get('nombre','Negocio'))}**", unsafe_allow_html=True)
        st.caption(email)
        st.divider()
        paginas = ["Inicio", "Caja diaria", "Ventas y gastos", "Clientes", "Proveedores", "Cuentas", "Inventario", "Reportes", "Alertas", "BradaFin IA", "Perfil"]
        iconos = {"Inicio":"🏠", "Caja diaria":"💵", "Ventas y gastos":"🧾", "Clientes":"👥", "Proveedores":"🚚", "Cuentas":"📌", "Inventario":"📦", "Reportes":"📊", "Alertas":"🔔", "BradaFin IA":"🤖", "Perfil":"⚙️"}
        for p in paginas:
            if st.button(f"{iconos.get(p,'•')} {p}", key=f"nav_{p}", use_container_width=True, type="primary" if st.session_state.pagina == p else "secondary"):
                st.session_state.pagina = p
                st.rerun()
        st.divider()
        st.caption("BradaFin · negocios. Zentix queda para personas.")

    mostrar_whatsapp_flash()
    pagina = st.session_state.pagina
    if pagina == "Inicio": render_inicio(negocio, user_id, df_movs, df_cuentas, df_productos)
    elif pagina == "Caja diaria": render_caja(negocio, user_id, df_movs, df_cajas)
    elif pagina == "Ventas y gastos": render_ventas_gastos(negocio, user_id, df_productos, df_movs)
    elif pagina == "Clientes": render_clientes(negocio, user_id, df_clientes, df_cuentas, df_abonos)
    elif pagina == "Proveedores": render_proveedores(negocio, user_id, df_proveedores)
    elif pagina == "Cuentas": render_cuentas(negocio, user_id, df_clientes, df_proveedores, df_cuentas, df_abonos)
    elif pagina == "Inventario": render_inventario(negocio, user_id, df_productos, df_movs)
    elif pagina == "Reportes": render_reportes(negocio, user_id, df_movs, df_cuentas, df_productos)
    elif pagina == "Alertas": render_alertas(negocio, user_id, df_movs, df_cuentas, df_productos)
    elif pagina == "BradaFin IA": render_ia(negocio, user_id, df_movs, df_cuentas, df_productos)
    elif pagina == "Perfil": render_perfil(negocio, user_id)
    st.markdown("<div class='muted' style='text-align:center;margin-top:1.5rem;'>BradaFin v1 · control, claridad y utilidad real para microempresas.</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
